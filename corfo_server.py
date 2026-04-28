"""
corfo_server.py  –  Flask API para el sistema CORFO Analytics
─────────────────────────────────────────────────────────────
Novedades respecto a la versión anterior:
  • Reemplaza google-genai por Mellea (IBM) como capa de generación.
  • Mellea conecta a Gemini 2.5 Pro mediante su endpoint OpenAI-compatible.
  • El patrón Instruct → Validate → Repair (IVR) de Mellea hace hasta 3
    intentos automáticos, rechazando cualquier respuesta que no cumpla los
    requirements programáticos (JSON válido, SELECT, limit razonable).
  • Si los 3 intentos fallan, se devuelve el mejor resultado parcial junto
    con el error, así el usuario ve qué pasó en la UI.
"""

import os, sys, json, re, logging, time, threading, io, math, secrets, functools, hmac, warnings, traceback
import sqlite3  # kept for SQLite fallback; always available in stdlib
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

warnings.filterwarnings("ignore", message=".*pandas only supports SQLAlchemy.*")

# ── Database driver selection ─────────────────────────────────────────────────
# If DATABASE_URL is set, connect via psycopg2 (PostgreSQL).
# Otherwise fall back to SQLite for local development.
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    _PSYCOPG2_AVAILABLE = True
except ImportError:
    _PSYCOPG2_AVAILABLE = False

# Unified operational-error tuple used in except clauses throughout the file.
_DB_OPERATIONAL_ERROR: tuple = (
    *(  (psycopg2.OperationalError, psycopg2.Error) if _PSYCOPG2_AVAILABLE else ()  ),
    sqlite3.OperationalError,
    sqlite3.Error,
)


def _resolve_database_url() -> str | None:
    # PG* individual vars take priority — Railway injects these reliably per-service.
    pguser = os.environ.get("PGUSER", "").strip()
    pghost = os.environ.get("PGHOST", "").strip()
    pgdatabase = os.environ.get("PGDATABASE", "").strip()
    if pguser and pghost and pgdatabase:
        pgpassword = os.environ.get("PGPASSWORD", "").strip()
        pgport = os.environ.get("PGPORT", "5432").strip()
        return f"postgresql://{pguser}:{pgpassword}@{pghost}:{pgport}/{pgdatabase}"
    return (
        os.environ.get("DATABASE_URL", "").strip()
        or None
    )


def is_postgres() -> bool:
    return bool(_resolve_database_url() and _PSYCOPG2_AVAILABLE)


def get_db():
    """Open a database connection (PostgreSQL or SQLite fallback)."""
    database_url = _resolve_database_url()
    if database_url and _PSYCOPG2_AVAILABLE:
        if database_url.startswith("postgres://"):
            database_url = database_url.replace("postgres://", "postgresql://", 1)
        conn = psycopg2.connect(database_url)
        return conn
    conn = sqlite3.connect(os.environ.get("DB_PATH", "corfo_alimentos.db"), timeout=10)
    conn.row_factory = sqlite3.Row
    return conn


def _init_postgres_schema() -> None:
    """Create all tables and indexes on first boot when using PostgreSQL.

    Reads 003_create_postgresql_schema.sql which uses CREATE TABLE IF NOT EXISTS,
    so it is safe to run on every startup.
    """
    if not is_postgres():
        return
    schema_path = Path(__file__).parent / "sync" / "schema_migrations" / "003_create_postgresql_schema.sql"
    if not schema_path.exists():
        log.warning("PostgreSQL schema file not found: %s", schema_path)
        return
    ddl = schema_path.read_text(encoding="utf-8")
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(ddl)
        conn.commit()
        log.info("PostgreSQL schema initialized from %s", schema_path.name)
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        log.info("PostgreSQL schema already exists (concurrent worker init) — OK")
    except psycopg2.errors.DuplicateTable:
        conn.rollback()
        log.info("PostgreSQL schema already exists — OK")
    except Exception as e:
        log.error("Failed to initialize PostgreSQL schema: %s", e)
        conn.rollback()
        raise
    finally:
        conn.close()


def get_cursor(conn):
    """Return a dict-capable cursor for the given connection.

    Uses RealDictCursor for PostgreSQL; standard cursor for SQLite
    (which already returns sqlite3.Row objects via row_factory).
    """
    if is_postgres():
        return conn.cursor(cursor_factory=RealDictCursor)
    return conn.cursor()


def _sql(query: str) -> str:
    """Adapt SQL placeholder style for the current driver.

    SQLite uses '?' while psycopg2 uses '%s'. All queries in this file
    are written with '?' and this helper translates them at call time.
    """
    if is_postgres():
        return query.replace("?", "%s")
    return query

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from sync.datainnovacion_sync import run_sync

load_dotenv()
from flask import Flask, request, jsonify, send_file, Response, session
from flask_cors import CORS
from werkzeug.security import check_password_hash, generate_password_hash
import pandas as pd

# ── Mellea ──────────────────────────────────────────────────────────────────
import mellea
from mellea import MelleaSession
from mellea.backends.openai import OpenAIBackend
from mellea.stdlib.sampling import RejectionSamplingStrategy
from mellea.core import Requirement
from mellea.stdlib.requirements import simple_validate
from apscheduler.schedulers.background import BackgroundScheduler

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("corfo")
audit_log = logging.getLogger("corfo.audit")

_QUERY_LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "query_log.jsonl")


def _log_query(prompt: str, sql: str | None, row_count: int | None, error: str | None) -> None:
    entry = {
        "timestamp": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "prompt": prompt,
        "sql": sql,
        "row_count": row_count,
        "error": error,
    }
    try:
        with open(_QUERY_LOG_PATH, "a", encoding="utf-8") as fh:
            fh.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except OSError as exc:
        log.warning("No se pudo escribir en query_log.jsonl: %s", exc)


# ─────────────────────────────────────────────────────────────────────────────
# RATE LIMITING
# ─────────────────────────────────────────────────────────────────────────────
RATE_LIMIT_QUERY_MAX    = 20    # requests
RATE_LIMIT_QUERY_WINDOW = 60    # seconds
RATE_LIMIT_SYNC_MAX     = 3     # requests
RATE_LIMIT_SYNC_WINDOW  = 3600  # seconds


class RateLimiter:
    """Sliding-window in-memory rate limiter. Thread-safe via Lock."""

    def __init__(self, max_requests: int, window_seconds: int) -> None:
        self._max = max_requests
        self._window = window_seconds
        self._store: dict[str, list[float]] = {}  # ip → [timestamp, ...]
        self._lock = threading.Lock()

    def is_allowed(self, ip: str) -> tuple[bool, int]:
        """Check whether the IP is within the rate limit.

        Returns (allowed, retry_after_seconds).
        retry_after_seconds is 0 when allowed.
        """
        now = time.monotonic()
        cutoff = now - self._window

        with self._lock:
            timestamps = self._store.get(ip, [])
            # Prune entries outside the sliding window
            timestamps = [t for t in timestamps if t > cutoff]

            if len(timestamps) >= self._max:
                # Oldest timestamp tells us when a slot frees up
                retry_after = int(timestamps[0] - cutoff) + 1
                self._store[ip] = timestamps
                return False, retry_after

            timestamps.append(now)
            self._store[ip] = timestamps
            return True, 0


_query_limiter = RateLimiter(RATE_LIMIT_QUERY_MAX, RATE_LIMIT_QUERY_WINDOW)
_sync_limiter  = RateLimiter(RATE_LIMIT_SYNC_MAX,  RATE_LIMIT_SYNC_WINDOW)


def _get_client_ip() -> str:
    """Return client IP, honouring X-Forwarded-For for proxy deployments."""
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.remote_addr or "unknown"


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
app  = Flask(__name__)

# Session security — SECRET_KEY is required for signed cookies.
# If not set in .env, generate a random key (non-persistent across restarts).
_secret_key = os.getenv("SECRET_KEY")
if not _secret_key:
    _secret_key = secrets.token_hex(32)
    log.warning("SECRET_KEY no está configurado en .env — se usó una clave temporal. Las sesiones no sobrevivirán reinicios.")
app.secret_key = _secret_key
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = not app.debug

# ── Credenciales — leídas desde .env ────────────────────────────────────────
# Usuario principal (admin por defecto)
_ADMIN_USERNAME: str = os.getenv("ADMIN_USERNAME", "")
_ADMIN_PASSWORD_HASH: str = os.getenv("ADMIN_PASSWORD", "")
_ADMIN_ROLE_RAW: str = os.getenv("ADMIN_ROLE", "admin").strip().lower()
_VALID_ROLES: set[str] = {"admin", "viewer"}
if _ADMIN_ROLE_RAW not in _VALID_ROLES:
    log.warning(
        "ADMIN_ROLE='%s' no es válido (se esperaba 'admin' o 'viewer'). "
        "Se usará 'admin' como valor por defecto.",
        _ADMIN_ROLE_RAW,
    )
    _ADMIN_ROLE_RAW = "admin"
_ADMIN_ROLE: str = _ADMIN_ROLE_RAW

# Usuario viewer opcional — VIEWER_USERNAME + VIEWER_PASSWORD en .env
# Si no están configurados, el acceso de viewer no está disponible.
_VIEWER_USERNAME: str = os.getenv("VIEWER_USERNAME", "")
_VIEWER_PASSWORD_HASH: str = os.getenv("VIEWER_PASSWORD", "")
# Pre-computed at startup so _get_db_user can do a constant-time dummy check
# when a username is not found, preventing timing-based username enumeration.
_DUMMY_HASH: str = generate_password_hash("__dummy__")


def _check_credentials(username: str, password: str) -> str | None:
    """Verifica username+password contra las credenciales en .env.

    Retorna el rol ('admin' o 'viewer') si las credenciales son válidas,
    o None si no coinciden con ningún usuario.
    """
    if _ADMIN_USERNAME and hmac.compare_digest(username, _ADMIN_USERNAME):
        if check_password_hash(_ADMIN_PASSWORD_HASH, password):
            return _ADMIN_ROLE
    if _VIEWER_USERNAME and hmac.compare_digest(username, _VIEWER_USERNAME):
        if check_password_hash(_VIEWER_PASSWORD_HASH, password):
            return "viewer"
    return None

CORS(app)

# Scheduler mensual — corre el día 1 de cada mes a las 03:00
# No se inicia bajo pytest para evitar efectos secundarios en tests.
if "pytest" not in sys.modules:
    _scheduler = BackgroundScheduler()
    _scheduler.add_job(run_sync, "cron", day=1, hour=3, minute=0, id="monthly_sync")
    _scheduler.start()
    log.info("APScheduler iniciado — sync mensual programado para el día 1 de cada mes a las 03:00")

DB   = os.environ.get("DB_PATH", os.path.join(os.path.dirname(os.path.abspath(__file__)), 'corfo_alimentos.db'))


MODEL_SQL     = "llama-3.3-70b-versatile"  # Groq via OpenAI-compat endpoint
MODEL_EXPLAIN = "llama-3.3-70b-versatile"
GEMINI_BASE_URL = "https://api.groq.com/openai/v1"
GEMINI_API_KEY  = os.getenv("GROQ_API_KEY")
MAX_RETRIES  = 3                  # loop_budget para RejectionSamplingStrategy
# Groq free-tier TPM is lower on 70B — retry 429s with exponential backoff.
GROQ_MAX_BACKOFF_RETRIES = 4     # up to ~60s total wait
GROQ_BACKOFF_BASE        = 5     # seconds — doubles each attempt

# ─────────────────────────────────────────────────────────────────────────────
# MELLEA SESSIONS  (lazy – se crean al primer uso)
# ─────────────────────────────────────────────────────────────────────────────
_sql_session: MelleaSession | None = None
_explain_session: MelleaSession | None = None

def _get_sql_session() -> MelleaSession:
    """Session para generar SQL (Gemini 2.0 Flash via OpenAI-compat)."""
    global _sql_session
    if _sql_session is not None:
        return _sql_session

    backend = OpenAIBackend(
        model_id=MODEL_SQL,
        base_url=GEMINI_BASE_URL,
        api_key=GEMINI_API_KEY,
        model_options={"max_tokens": 1024},
        default_to_constraint_checking_alora=False,
    )
    _sql_session = MelleaSession(backend=backend)
    log.info("SQL session creada -> %s via Groq", MODEL_SQL)
    return _sql_session

def _get_explain_session() -> MelleaSession:
    """Session para explicar resultados (Gemini 2.0 Flash via OpenAI-compat)."""
    global _explain_session
    if _explain_session is not None:
        return _explain_session

    backend = OpenAIBackend(
        model_id=MODEL_EXPLAIN,
        base_url=GEMINI_BASE_URL,
        api_key=GEMINI_API_KEY,
        model_options={"max_tokens": 512},
        default_to_constraint_checking_alora=False,
    )
    _explain_session = MelleaSession(backend=backend)
    log.info("Explain session creada -> %s via Groq", MODEL_EXPLAIN)
    return _explain_session

# ─────────────────────────────────────────────────────────────────────────────
# SCHEMA CONTEXT  (se construye una vez al arrancar)
# ─────────────────────────────────────────────────────────────────────────────
# Columnas categóricas de las que se enumeran valores distintos en el contexto.
# Sólo columnas con cardinalidad baja-media y relevancia para el SQL generado.
_CATEGORICAL_COLS: list[str] = [
    "foco_apoyo",
    "tipo_intervencion",
    "instrumento_homologado",
    "estado_data",
    "tipo_innovacion",
    "region_ejecucion",
    "tramo_ventas",
    "tendencia_final",
    "tipo_proyecto",
    "sostenible",
    "economia_circular_si_no",
]
_DISTINCT_LIMIT = 20      # max distinct values shown per column
_VALUE_MAX_CHARS = 60     # max characters per value string


def _list_table_columns(cur, table: str) -> list[tuple[str, str]]:
    """Return (column_name, type) pairs for the given table.

    Uses PRAGMA on SQLite and information_schema on PostgreSQL.
    """
    if is_postgres():
        cur.execute(
            """
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_name = %s
            ORDER BY ordinal_position
            """,
            (table,),
        )
        return [(row["column_name"], row["data_type"]) for row in cur.fetchall()]
    else:
        cur.execute(f"PRAGMA table_info({table})")
        return [(row[1], row[2]) for row in cur.fetchall()]


def _build_schema_context() -> str:
    conn = get_db()
    cur = get_cursor(conn)
    lines = []

    lines.append("=== TABLE: proyectos ===")
    for col_name, col_type in _list_table_columns(cur, "proyectos"):
        lines.append(f"  {col_name} ({col_type})")

    # Enumerate distinct values for selected categorical columns.
    # Cap at _DISTINCT_LIMIT values; truncate each value to _VALUE_MAX_CHARS chars.
    lines.append("  KEY VALUES:")
    lines.append("  sector_economico variants include: 'Agrícola excepto vitivinícola', 'Alimentos excepto vitivinícola' — use LIKE '%Agrícola%' or '%Alimentos%' for broad matches.")
    if is_postgres():
        lines.append("  aprobado_corfo: NUMERIC (e.g. 19042000). Use directly or CAST(aprobado_corfo AS FLOAT).")
    else:
        lines.append("  aprobado_corfo: TEXT, e.g. '19042000'. Cast: CAST(aprobado_corfo AS REAL).")
    lines.append('  año_adjudicacion: INTEGER 2009-2025. Always double-quote: "año_adjudicacion".')

    for col in _CATEGORICAL_COLS:
        try:
            # Fetch one extra row to detect whether the full list was cut
            cur.execute(
                _sql(
                    f'SELECT DISTINCT "{col}" FROM proyectos '
                    f'WHERE "{col}" IS NOT NULL '
                    f'ORDER BY "{col}" LIMIT ?'
                ),
                (_DISTINCT_LIMIT + 1,),
            )
            rows = cur.fetchall()
        except Exception:
            if is_postgres():
                conn.rollback()
            continue  # skip if column doesn't exist

        total_fetched = len(rows)
        capped = rows[:_DISTINCT_LIMIT]

        # Truncate each value string to _VALUE_MAX_CHARS characters
        values: list[str] = []
        for row in capped:
            v = row[0] if not isinstance(row, dict) else row[col]
            s = str(v)
            if len(s) > _VALUE_MAX_CHARS:
                s = s[:_VALUE_MAX_CHARS] + "…"
            values.append(repr(s))

        suffix = ""
        if total_fetched > _DISTINCT_LIMIT:
            # Count actual total to show how many more exist
            cur.execute(
                f'SELECT COUNT(DISTINCT "{col}") AS count FROM proyectos WHERE "{col}" IS NOT NULL'
            )
            row = cur.fetchone()
            total_count = row[0] if not isinstance(row, dict) else row["count"]
            extra = total_count - _DISTINCT_LIMIT
            suffix = f" ... y {extra} más"

        lines.append(f"  {col}: {', '.join(values)}{suffix}")

    lines.append("\n=== TABLE: leads ===")
    for col_name, col_type in _list_table_columns(cur, "leads"):
        lines.append(f"  {col_name} ({col_type})")

    conn.close()
    return "\n".join(lines)

_init_postgres_schema()
SCHEMA_CONTEXT = _build_schema_context()
log.info("SCHEMA_CONTEXT tamaño: %d chars, ~%d tokens aprox.", len(SCHEMA_CONTEXT), len(SCHEMA_CONTEXT) // 4)


def _build_context_enrichment() -> str:
    """Carga query_examples.json y las reglas críticas de field_dictionary.json,
    además de query_notes completos y significados de negocio desde schema_context.md."""
    lines = []
    base = os.path.dirname(os.path.abspath(__file__))

    # Few-shot examples
    examples_path = os.path.join(base, "context", "query_examples.json")
    try:
        with open(examples_path, encoding="utf-8") as f:
            examples = json.load(f)
        lines.append("FEW-SHOT EXAMPLES (question → correct SQL):")
        for ex in examples[:10]:  # primeros 10 para no sobrecargar el prompt
            lines.append(f"  Q: {ex['question_es']}")
            lines.append(f"  SQL: {ex['sql']}")
            if ex.get("notes"):
                lines.append(f"  Note: {ex['notes']}")
            lines.append("")
    except Exception as e:
        log.warning("No se pudo cargar query_examples.json: %s", e)

    # field_dictionary.json — dos bloques: quirks críticos + query_notes de proyectos
    dict_path = os.path.join(base, "context", "field_dictionary.json")
    try:
        with open(dict_path, encoding="utf-8") as f:
            fdict = json.load(f)

        # Bloque 1: quirks marcados como CRÍTICO (comportamiento existente, sin cambios)
        lines.append("CRITICAL FIELD QUIRKS:")
        for table, fields in fdict.items():
            if table.startswith("_"):
                continue
            for field, meta in fields.items():
                if meta.get("quirks") and "CRÍTICO" in meta.get("quirks", ""):
                    lines.append(f"  {field}: {meta['quirks']}")

        # Bloque 2: query_notes de todos los campos de proyectos con valor no vacío
        proyectos_fields = fdict.get("proyectos", {})
        notes_lines = []
        for field, meta in proyectos_fields.items():
            qn = meta.get("query_notes", "").strip()
            if qn:
                notes_lines.append(f"  {field}: {qn}")
        if notes_lines:
            lines.append("")
            lines.append("FIELD QUERY NOTES (proyectos):")
            lines.extend(notes_lines)

    except Exception as e:
        log.warning("No se pudo cargar field_dictionary.json: %s", e)

    # schema_context.md — bloque compacto de significados de negocio
    # El archivo tiene una tabla markdown en la sección "## Tabla: proyectos"
    # con filas: | `campo` | nombre usuario | tipo | notas |
    # Se extraen las filas de datos (no encabezado ni separador) para los campos
    # priorizados. Si hay espacio, se añaden los demás campos de la tabla.
    schema_path = os.path.join(base, "context", "schema_context.md")
    _PRIORITY_FIELDS = {
        "tendencia_final", "sector_economico", "tipo_proyecto", "tipo_innovacion",
        "sostenible", "economia_circular_si_no", "region_ejecucion",
        "año_adjudicacion", "aprobado_corfo",
    }
    try:
        with open(schema_path, encoding="utf-8") as f:
            md_lines = f.readlines()

        # Collect table rows under "## Tabla: proyectos" until next "##" heading
        in_proyectos = False
        table_rows: list[tuple[str, str]] = []  # (field_name, notes_text)
        for raw in md_lines:
            stripped = raw.strip()
            if stripped.startswith("## Tabla: proyectos"):
                in_proyectos = True
                continue
            if in_proyectos and stripped.startswith("##"):
                break
            if not in_proyectos:
                continue
            # Table data row: starts and ends with |, not a separator row (--|)
            if stripped.startswith("|") and "--" not in stripped:
                cells = [c.strip() for c in stripped.split("|") if c.strip()]
                if len(cells) < 4:
                    continue  # header row or malformed
                raw_field = cells[0].strip("`")
                # Skip the header row (field cell would be "Campo" or similar)
                if raw_field.lower() in ("campo", "field"):
                    continue
                business_notes = cells[3]  # 4th column = "Notas técnicas"
                if raw_field and business_notes:
                    table_rows.append((raw_field, business_notes))

        # Separate priority vs. other rows, cap total at 30 lines
        priority_rows = [(f, n) for f, n in table_rows if f in _PRIORITY_FIELDS]
        other_rows    = [(f, n) for f, n in table_rows if f not in _PRIORITY_FIELDS]
        selected = priority_rows + other_rows
        selected = selected[:30]

        if selected:
            lines.append("")
            lines.append("FIELD BUSINESS MEANINGS (proyectos):")
            for field, note in selected:
                lines.append(f"  {field}: {note}")

    except Exception as e:
        log.warning("No se pudo cargar schema_context.md: %s", e)

    return "\n".join(lines)


CONTEXT_ENRICHMENT = _build_context_enrichment()


# ─────────────────────────────────────────────────────────────────────────────
# FEW-SHOT EXAMPLES  (pares pregunta→SQL validados para guiar al modelo)
# ─────────────────────────────────────────────────────────────────────────────
def _build_few_shot_examples() -> str:
    """Carga query_examples.json y devuelve un subconjunto representativo
    formateado como bloques Q/SQL. Retorna "" si el archivo no existe."""
    base = os.path.dirname(os.path.abspath(__file__))
    examples_path = os.path.join(base, "context", "query_examples.json")
    # Índices seleccionados por diversidad de patrones SQL
    SELECTED_INDICES = [0, 4, 10]
    try:
        with open(examples_path, encoding="utf-8") as f:
            examples = json.load(f)
        lines: list[str] = []
        for i in SELECTED_INDICES:
            if i < len(examples):
                ex = examples[i]
                lines.append(f"Q: {ex['question_es']}")
                lines.append(f"SQL: {ex['sql']}")
                lines.append("")
        return "\n".join(lines).rstrip()
    except Exception as e:
        log.warning("No se pudo cargar query_examples.json para few-shot: %s", e)
        return ""


FEW_SHOT_EXAMPLES = _build_few_shot_examples()

# ─────────────────────────────────────────────────────────────────────────────
# FIELD DICTIONARY  (cargado una vez al arrancar para recuperación dinámica)
# ─────────────────────────────────────────────────────────────────────────────
def _load_field_dict() -> dict:
    """Carga field_dictionary.json una vez al arrancar. Retorna {} si falla."""
    base = os.path.dirname(os.path.abspath(__file__))
    dict_path = os.path.join(base, "context", "field_dictionary.json")
    try:
        with open(dict_path, encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        log.warning("No se pudo cargar field_dictionary.json para recuperación dinámica: %s", e)
        return {}

_FIELD_DICT: dict = _load_field_dict()

# ─────────────────────────────────────────────────────────────────────────────
# PROMPT TEMPLATE  (se envía como instrucción a Mellea)
# ─────────────────────────────────────────────────────────────────────────────
SQL_INSTRUCTION_TEMPLATE = (
    "You are a SQL expert for a Chilean public database of R&D and innovation projects funded by CORFO (Corporación de Fomento de la Producción). The database covers projects across all economic sectors in Chile.\n"
    "Given the user question, produce a JSON response.\n\n"
    "Respond with ONLY a valid JSON object — no markdown, no fences, nothing else:\n"
    '{% raw %}{\n'
    '  "sql": "<valid PostgreSQL SELECT or null>",\n'
    '  "chart_type": "<bar|line|pie|null>"\n'
    '}{% endraw %}\n\n'
    "RULES:\n"
    "- Only SELECT. Never INSERT/UPDATE/DELETE/DROP.\n"
    '- Double-quote columns with special chars: "año_adjudicacion".\n'
    "- aprobado_corfo is NUMERIC — use directly or CAST(aprobado_corfo AS FLOAT). No TEXT cast needed.\n"
    "- LIMIT 50 unless the user explicitly asks for more.\n"
    "- If you cannot answer, set sql to null.\n"
    "- PostgreSQL strict GROUP BY: every non-aggregated column in SELECT must appear in GROUP BY. Never select a column that is not grouped or aggregated.\n"
    "- Use BIGINT for large sums: SUM(aprobado_corfo)::BIGINT. Never cast sums to INTEGER — Chilean peso amounts overflow 32-bit integers.\n"
    '- "año_adjudicacion" is INTEGER — never use LIKE on it. Filter with = or BETWEEN: "año_adjudicacion" = 2024 or "año_adjudicacion" BETWEEN 2020 AND 2024.\n'
    "- Never alias non-company columns as 'razon' or 'empresa'. Only use those names when selecting the actual razon column.\n"
    "- Always exclude razon = 'Persona Natural' from company rankings, listings, and GROUP BY queries. These are anonymized individuals, not real companies.\n"
    "- When the question asks about proyectos/iniciativas/programas, always SELECT at minimum: codigo, instrumento, razon, titulo_del_proyecto, objetivo_general_del_proyecto, \"año_adjudicacion\".\n"
    "- When the question asks about empresas/compañías/razones sociales, GROUP BY razon and include COUNT(codigo) as cantidad_proyectos and SUM(aprobado_corfo)::BIGINT as monto_total_aprobado. ORDER BY monto_total_aprobado DESC.\n"
    "- SEMANTIC SEARCH (fallback): If the question contains <!-- semantic_keywords: k1, k2, ... -->, "
    "add LIKE conditions for each keyword across titulo_del_proyecto AND objetivo_general_del_proyecto "
    "joined with OR and wrapped in parentheses: (titulo_del_proyecto LIKE '%k1%' OR objetivo_general_del_proyecto LIKE '%k1%'). "
    "Do NOT change what columns to SELECT — follow the other rules for that.\n"
    "- SEMANTIC IDS: If the question contains <!-- semantic_ids: id1,id2,... -->, use a UNION to combine "
    "two result sets: (1) your primary SQL query with all user filters applied normally (no codigo restriction), "
    "and (2) a secondary SELECT for projects in the semantic IDs that were NOT already returned by the primary query. "
    "Structure it as:\n"
    "{% raw %}"
    "  SELECT <cols>, 0 AS _boost FROM proyectos WHERE <your_conditions>\n"
    "  UNION ALL\n"
    "  SELECT <cols>, 1 AS _boost FROM proyectos\n"
    "  WHERE codigo IN ('id1','id2',...)\n"
    "    AND codigo NOT IN (SELECT codigo FROM proyectos WHERE <your_conditions>)\n"
    "  ORDER BY _boost ASC, <your_order>\n"
    "  LIMIT 50"
    "{% endraw %}\n"
    "This ensures SQL precision is preserved while semantic results fill in any gaps. "
    "The LIMIT must appear exactly once, after the final ORDER BY, applying to the entire UNION result — never inside individual SELECT branches. "
    "If the primary query has no meaningful WHERE conditions (e.g., no filters at all), use ONLY the semantic IDs "
    "via `codigo IN (...)` to avoid returning the entire table. "
    "Do NOT add LIKE conditions when semantic IDs are present. "
    "Do NOT change what columns to SELECT — follow the proyectos/empresas rules below for that.\n"
    "- COUNT vs SUM: 'cuántos proyectos/iniciativas/programas' → COUNT(*). "
    "'cuánto monto/dinero/financiamiento/se aprobó/se financió' → SUM(aprobado_corfo)::BIGINT. "
    "Never use SUM to count records, never use COUNT for money totals.\n"
    "- DEFAULT ORDER: Unless the question specifies a sort order, add "
    'ORDER BY "año_adjudicacion" DESC to project listings (no GROUP BY) so the most recent appear first. '
    'Exception: time-series queries (GROUP BY "año_adjudicacion") use ORDER BY "año_adjudicacion" ASC. '
    'For all other GROUP BY queries, ORDER BY an aggregated column (e.g. COUNT or SUM result), never by a column not in SELECT.\n\n'
    "RAZONAMIENTO PREVIO AL SQL:\n"
    "Antes de escribir el SQL, identifica internamente:\n"
    "  (a) ¿Qué entidad cuenta o suma el usuario? (proyectos, empresas, montos)\n"
    "  (b) ¿Qué filtros aplica? (región, sector, año, tipo, etc.)\n"
    "  (c) ¿Necesita agrupar, comparar entre períodos, calcular porcentajes o hacer ranking?\n"
    "Luego escribe el SQL más preciso posible para esa intención.\n\n"
    + (
        "EXAMPLES:\n"
        f"{FEW_SHOT_EXAMPLES}\n\n"
        if FEW_SHOT_EXAMPLES else ""
    ) +
    "DATABASE SCHEMA:\n"
    f"{SCHEMA_CONTEXT}\n\n"
    "User question: {{question}}"
)

EXPLAIN_INSTRUCTION_TEMPLATE = (
    "Eres un analista experto en proyectos CORFO del sector alimentario chileno.\n"
    "El usuario hizo esta pregunta: {{question}}\n\n"
    "Se ejecutó esta consulta SQL: {{sql}}\n\n"
    "Y se obtuvieron estos resultados (máx 10 filas de muestra):\n{{results}}\n\n"
    "Total de filas: {{total_rows}}\n\n"
    "Tu tarea: Escribe una respuesta en español de 2-4 oraciones que:\n"
    "1. Responda directamente la pregunta del usuario\n"
    "2. Resuma los hallazgos principales de los datos\n"
    "3. Mencione cifras o tendencias relevantes si las hay\n\n"
    "Responde SOLO con el texto de la explicación, sin formato JSON ni markdown."
)

# ─────────────────────────────────────────────────────────────────────────────
# VALIDATION HELPERS  (used as Mellea requirements)
# ─────────────────────────────────────────────────────────────────────────────
def _is_valid_json(text: str) -> bool:
    """Requirement 1: la respuesta debe ser JSON parseable."""
    try:
        json.loads(_strip_fences(text))
        return True
    except (json.JSONDecodeError, ValueError):
        return False

def _is_select_or_null(text: str) -> bool:
    """Requirement 2: si hay SQL debe ser un SELECT (o null)."""
    try:
        obj = json.loads(_strip_fences(text))
        sql = obj.get("sql")
        if sql is None:
            return True                          # null es válido
        return sql.strip().upper().startswith("SELECT")
    except Exception:
        return False                             # si no parsea, dejamos que req 1 lo reporte

def _strip_fences(text: str) -> str:
    """Extract the first JSON object from model output, ignoring preamble/postamble.

    Handles: bare JSON, ```json fences, and verbose models that add text before/after.
    """
    text = text.strip()
    # Fast path: already valid JSON
    try:
        json.loads(text)
        return text
    except (json.JSONDecodeError, ValueError):
        pass
    # Strip code fences
    text = re.sub(r'^```(?:json)?\s*', '', text)
    text = re.sub(r'\s*```$', '', text).strip()
    try:
        json.loads(text)
        return text
    except (json.JSONDecodeError, ValueError):
        pass
    # Extract first {...} block — handles preamble text from verbose models
    match = re.search(r'\{[\s\S]*\}', text)
    if match:
        return match.group(0)
    return text


# ── Validación de calidad del resultado ──────────────────────────────────────

_BROAD_QUERY_WORDS = {"todos", "todo", "all", "completo", "completa", "todas"}

def _check_result_quality(question: str, df: pd.DataFrame) -> str | None:
    """
    Detecta resultados vacíos, sospechosamente triviales o excesivamente grandes.

    Retorna una advertencia en español, o None si el resultado parece correcto.
    """
    # 1) Resultado vacío
    if df.empty:
        log.debug("Resultado vacío para la pregunta: %s", question)
        return "La consulta no devolvió resultados. Intenta reformular la pregunta."

    # 2) Celda única con valor 0 o None
    if df.shape == (1, 1):
        valor = df.iat[0, 0]
        if valor is None or valor == 0:
            log.debug("Resultado de celda única sospechoso (%r) para: %s", valor, question)
            return "La consulta no devolvió resultados. Intenta reformular la pregunta."

    # 3) Resultado grande sin palabras que indiquen intención de listar todo.
    # Se usa > 50 (no >=) porque el prompt instruye LIMIT 50 por defecto,
    # por lo que exactamente 50 filas es el resultado normal esperado.
    if len(df) > 50:
        pregunta_lower = question.lower()
        if not any(palabra in pregunta_lower for palabra in _BROAD_QUERY_WORDS):
            log.debug("Resultado grande (%d filas) para pregunta sin 'todos/todo/all': %s", len(df), question)
            return "Se devolvieron muchos resultados. Considera filtrar por año, región o sector."

    return None


def _is_empty_result_warning(warning: str | None) -> bool:
    """
    Devuelve True cuando la advertencia indica resultado vacío o sospechoso
    (casos 1 y 2 de _check_result_quality), pero NO el caso de muchos resultados.
    """
    return warning is not None and "muchos resultados" not in warning


_FALLBACK_PREAMBLES: dict[str, str] = {
    "no_sql":             "No pude generar una consulta para esa pregunta.",
    "low_quality_result": "La consulta no devolvió resultados útiles.",
    "hallucinated_column":"La consulta generada referenciaba campos que no existen.",
    "ivr_exhausted":      "El modelo no logró generar una consulta válida tras varios intentos.",
}

_FALLBACK_SUGGESTIONS = (
    "Para obtener mejores resultados, considera reformular la pregunta:\n"
    "- Usa nombres de sector específicos (por ejemplo: \"Alimentos\", \"Biotecnología\", \"TIC\").\n"
    "- Especifica un rango de años (por ejemplo: \"entre 2018 y 2023\").\n"
    "- Menciona una región por su nombre oficial (por ejemplo: \"Región Metropolitana\", \"Biobío\").\n\n"
    "Ejemplo de pregunta bien formulada: "
    "\"¿Cuáles son las empresas con más financiamiento CORFO en el sector Alimentos entre 2019 y 2022?\""
)


def _build_fallback_response(question: str, reason: str) -> dict:
    """
    Construye una respuesta de baja confianza lista para jsonify().
    Función pura: sin llamadas a BD ni al modelo.
    """
    preamble = _FALLBACK_PREAMBLES.get(reason, "No se pudo responder la pregunta con confianza.")
    answer = f"{preamble}\n\n{_FALLBACK_SUGGESTIONS}"
    return {
        "answer":         answer,
        "dataframe":      [],
        "chart_type":     None,
        "sql":            None,
        "low_confidence": True,
        "fallback_reason": reason,
        "warning":        preamble,
    }


# ─────────────────────────────────────────────────────────────────────────────
# RECUPERACIÓN DINÁMICA DE CAMPOS RELEVANTES
# ─────────────────────────────────────────────────────────────────────────────
def _retrieve_relevant_fields(question: str) -> str:
    """Identifica los campos de proyectos más relevantes para la pregunta del usuario.

    Usa coincidencia de cadenas contra _FIELD_DICT (sin LLM, sin BD).
    Retorna un bloque formateado con los campos de mayor puntuación, o "" si
    _FIELD_DICT no está disponible o ningún campo puntúa.
    """
    if not _FIELD_DICT:
        return ""

    proyectos_fields: dict = _FIELD_DICT.get("proyectos", {})
    if not proyectos_fields:
        return ""

    # Tokenizar la pregunta: palabras en minúsculas con longitud mínima 3
    question_lower = question.lower()
    tokens = {w for w in re.split(r'\W+', question_lower) if len(w) >= 3}

    scored: list[tuple[int, str, dict]] = []  # (score, field_name, meta)
    for field, meta in proyectos_fields.items():
        score = 0

        # +2 si el nombre del campo (completo o alguna parte separada por _) aparece en tokens
        field_parts = field.lower().split('_')
        if field.lower() in tokens or any(p in tokens for p in field_parts if len(p) >= 3):
            score += 2

        # +1 por cada example_value cuya forma en minúsculas aparece como subcadena
        for example in meta.get("example_values", []):
            if str(example).lower() in question_lower:
                score += 1

        # +1 si alguna palabra de description_es aparece en los tokens
        description = meta.get("description_es", "")
        desc_words = {w for w in re.split(r'\W+', description.lower()) if len(w) >= 3}
        if desc_words & tokens:
            score += 1

        if score > 0:
            scored.append((score, field, meta))

    if not scored:
        return ""

    # Ordenar por puntuación descendente, tomar top 8
    scored.sort(key=lambda x: x[0], reverse=True)
    top_fields = scored[:8]

    lines = ["CAMPOS RELEVANTES PARA ESTA CONSULTA:"]
    for _, field, meta in top_fields:
        display = meta.get("display_name_es", field)
        notes   = meta.get("query_notes", "").strip()
        lines.append(f"  {field}: {display} — {notes}")

    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# EXPANSIÓN SEMÁNTICA DE PALABRAS CLAVE
# ─────────────────────────────────────────────────────────────────────────────

# Valores estructurados conocidos — si la pregunta menciona alguno, es una
# consulta estructurada y no requiere expansión semántica.
_KNOWN_STRUCTURED_VALUES: set[str] = {
    # años
    "2009", "2010", "2011", "2012", "2013", "2014", "2015", "2016",
    "2017", "2018", "2019", "2020", "2021", "2022", "2023", "2024", "2025",
    # regiones (parcial — coincidencia con subcadena en la pregunta)
    "metropolitana", "valparaíso", "biobío", "araucanía", "maule",
    "coquimbo", "antofagasta", "atacama", "aysén", "magallanes",
    "tarapacá", "arica", "ñuble", "los ríos", "los lagos",
    # sectores conocidos
    "alimentos", "biotecnología", "acuicultura", "pesca", "silvoagropecuario",
    "tecnologías de la información", "minería", "energía", "salud",
    # tendencias conocidas
    "inteligencia artificial", "internet de las cosas", "drones", "genómica",
    "nanotecnología", "biotecnología", "gamificación",
    # campos booleanos
    "sostenible", "economía circular", "vigente", "finalizado",
}

# Indicadores de que la pregunta pide búsqueda semántica por concepto
_SEMANTIC_TRIGGER_WORDS: set[str] = {
    "desarrollando", "trabajan", "desarrollan", "fabrican", "producen",
    "trabaja", "fabrica", "produce", "enfocadas", "orientadas", "dedicadas",
    "especializadas", "innovan", "investigan",
}

_KEYWORD_EXPANSION_TEMPLATE = (
    "Eres un experto en industria alimentaria y tecnología chilena.\n"
    "El usuario busca proyectos relacionados con el concepto: {{concept}}\n\n"
    "Genera entre 5 y 8 palabras clave en español (y algún término técnico en inglés si aplica) "
    "que aparecerían en el título o descripción de ese tipo de proyecto. "
    "Incluye sinónimos, términos técnicos y variantes del concepto.\n\n"
    "Responde ÚNICAMENTE con una lista separada por comas, sin explicaciones, sin puntos, "
    "sin numeración. Ejemplo: snack, colación, saludable, funcional, nutritivo, aperitivo"
)


def _needs_semantic_expansion(question: str) -> bool:
    """Heurística para decidir si la pregunta requiere expansión semántica.

    Retorna True cuando:
    - La pregunta usa verbos que implican búsqueda conceptual (desarrollando, trabajan…), Y
    - No contiene valores estructurados conocidos (año, región, sector, tendencia).

    Usa coincidencia de palabra completa (\b) para valores estructurados, evitando
    que "salud" coincida dentro de "saludable".
    """
    q_lower = question.lower()

    # Coincidencia de palabra completa para valores estructurados
    if any(re.search(r'\b' + re.escape(val) + r'\b', q_lower)
           for val in _KNOWN_STRUCTURED_VALUES):
        return False

    # Si contiene algún verbo semántico → expandir
    return any(trigger in q_lower for trigger in _SEMANTIC_TRIGGER_WORDS)


def _extract_concept(question: str) -> str:
    """Extrae el concepto o término de búsqueda principal de la pregunta.

    Estrategia simple: toma la subcadena que sigue a la última preposición
    o verbo semántico conocido. Si no hay patrón claro, devuelve la pregunta completa.
    """
    q = question.strip().rstrip("?")
    # Buscar después de "en" / "sobre" / "de" seguido de sustantivo
    for marker in ("desarrollando ", "trabajan en ", "trabaja en ",
                   "desarrollan ", "fabrican ", "producen ",
                   "enfocadas en ", "orientadas a ", "sobre ", "de "):
        idx = q.lower().find(marker)
        if idx != -1:
            return q[idx + len(marker):].strip()
    return q


def _expand_keywords(question: str) -> list[str] | None:
    """Usa la sesión de explicación para expandir la pregunta con palabras clave semánticas.

    Retorna una lista de 5-8 keywords, o None si la pregunta no necesita expansión.
    """
    if not _needs_semantic_expansion(question):
        return None

    concept = _extract_concept(question)
    if not concept or len(concept) < 3:
        return None

    try:
        session = _get_explain_session()
        result = session.instruct(
            _KEYWORD_EXPANSION_TEMPLATE,
            user_variables={"concept": concept},
        )
        raw = str(result).strip()
        # Limpiar posibles artefactos y dividir por comas
        raw = _strip_fences(raw)
        keywords = [kw.strip().lower() for kw in raw.split(",") if kw.strip()]
        # Filtrar tokens vacíos o demasiado cortos
        keywords = [kw for kw in keywords if len(kw) >= 3]
        if keywords:
            log.info("Expansión semántica para '%s': %s", concept, keywords)
            return keywords[:8]
    except Exception as e:
        log.warning("_expand_keywords falló para '%s': %s", question, e)

    return None


# ─────────────────────────────────────────────────────────────────────────────
# BÚSQUEDA SEMÁNTICA POR EMBEDDINGS
# ─────────────────────────────────────────────────────────────────────────────

_embed_model = None
_embed_model_failed = False  # prevent repeated download attempts after a failure
_embed_rebuild_lock = threading.Lock()  # prevent concurrent rebuilds


def _get_embed_model():
    """Carga el modelo de embeddings de forma lazy (solo al primer uso)."""
    global _embed_model, _embed_model_failed
    if _embed_model_failed:
        return None
    if _embed_model is None:
        try:
            from sentence_transformers import SentenceTransformer
            log.info("Cargando modelo de embeddings paraphrase-multilingual-MiniLM-L12-v2 ...")
            _embed_model = SentenceTransformer("paraphrase-multilingual-MiniLM-L12-v2")
            log.info("Modelo de embeddings cargado OK.")
        except ImportError:
            log.warning(
                "sentence-transformers no está instalado — "
                "la búsqueda semántica por embeddings no estará disponible."
            )
            _embed_model_failed = True
        except Exception as e:
            log.warning(
                "Error cargando modelo de embeddings (%s) — búsqueda semántica desactivada.", e
            )
            _embed_model_failed = True
    return _embed_model


def _rebuild_embeddings_bg() -> None:
    """Rebuild proyectos_vec in-process using the already-loaded embed model.

    Runs in a background thread so the caller (sync or admin endpoint) returns
    immediately. Uses a lock to prevent concurrent rebuilds.
    """
    import numpy as np

    if not _embed_rebuild_lock.acquire(blocking=False):
        log.info("Rebuild de embeddings ya en curso — omitiendo.")
        return

    try:
        model = _get_embed_model()
        if model is None:
            log.warning("Rebuild embeddings: modelo no disponible — omitiendo.")
            return

        conn = get_db()
        cur = get_cursor(conn)
        cur.execute(
            "SELECT codigo, titulo_del_proyecto, objetivo_general_del_proyecto FROM proyectos"
        )
        rows = cur.fetchall()
        conn.close()

        if not rows:
            log.warning("Rebuild embeddings: tabla proyectos vacía — omitiendo.")
            return

        if isinstance(rows[0], dict):
            codigos = [r["codigo"] for r in rows]
            texts = [f"{r['titulo_del_proyecto'] or ''} {r['objetivo_general_del_proyecto'] or ''}".strip() for r in rows]
        else:
            codigos = [r[0] for r in rows]
            texts = [f"{r[1] or ''} {r[2] or ''}".strip() for r in rows]

        total = len(codigos)
        log.info("Rebuild embeddings: generando vectores para %d proyectos ...", total)
        embeddings = model.encode(texts, batch_size=64, show_progress_bar=False, convert_to_numpy=True)

        conn = get_db()
        cur = get_cursor(conn)
        if is_postgres():
            import psycopg2
            cur.execute("""
                CREATE TABLE IF NOT EXISTS proyectos_vec (
                    codigo TEXT PRIMARY KEY,
                    vector BYTEA NOT NULL
                )
            """)
            cur.execute("DELETE FROM proyectos_vec")
            for codigo, vec in zip(codigos, embeddings):
                blob = vec.astype("float32").tobytes()
                cur.execute(
                    "INSERT INTO proyectos_vec (codigo, vector) VALUES (%s, %s) "
                    "ON CONFLICT (codigo) DO UPDATE SET vector = EXCLUDED.vector",
                    (codigo, psycopg2.Binary(blob)),
                )
        else:
            cur.execute("DROP TABLE IF EXISTS proyectos_vec")
            cur.execute("CREATE TABLE proyectos_vec (codigo TEXT PRIMARY KEY, vector BLOB NOT NULL)")
            for codigo, vec in zip(codigos, embeddings):
                blob = vec.astype("float32").tobytes()
                cur.execute("INSERT INTO proyectos_vec (codigo, vector) VALUES (?, ?)", (codigo, blob))

        conn.commit()
        conn.close()
        log.info("Rebuild embeddings completado: %d vectores indexados.", total)
    except Exception as e:
        log.error("Rebuild embeddings falló: %s", e)
    finally:
        _embed_rebuild_lock.release()


def _trigger_embeddings_rebuild() -> None:
    """Launch _rebuild_embeddings_bg in a daemon thread."""
    t = threading.Thread(target=_rebuild_embeddings_bg, daemon=True, name="embed-rebuild")
    t.start()
    log.info("Rebuild de embeddings lanzado en background.")


def _semantic_ids(question: str, top_n: int = 50) -> list[str] | None:
    """Retorna los codigos de los top_n proyectos más similares semánticamente a la pregunta.

    Retorna None si la tabla proyectos_vec no existe, está vacía, o el modelo
    no está disponible (degradación elegante).
    """
    import numpy as np

    model = _get_embed_model()
    if model is None:
        return None

    try:
        conn = get_db()
        cur = get_cursor(conn)
        cur.execute("SELECT codigo, vector FROM proyectos_vec")
        rows = cur.fetchall()
        conn.close()
    except Exception as e:
        log.debug("proyectos_vec no disponible (%s) — omitiendo búsqueda semántica", e)
        return None

    if not rows:
        return None

    # Deserializar vectores — support both tuple rows (sqlite3) and dict rows (psycopg2)
    if rows and isinstance(rows[0], dict):
        ids = [r["codigo"] for r in rows]
        matrix = np.stack([np.frombuffer(bytes(r["vector"]), dtype=np.float32) for r in rows])
    else:
        ids = [r[0] for r in rows]
        matrix = np.stack([np.frombuffer(r[1], dtype=np.float32) for r in rows])

    # Embedding de la pregunta
    q_vec = model.encode(question, convert_to_numpy=True).astype(np.float32)

    # Similitud coseno
    norms = np.linalg.norm(matrix, axis=1)
    q_norm = np.linalg.norm(q_vec)
    if q_norm == 0:
        return None

    similarities = (matrix @ q_vec) / (norms * q_norm + 1e-10)

    # Top-N índices por similitud descendente
    top_indices = np.argsort(similarities)[::-1][:top_n]
    return [ids[i] for i in top_indices]


# Palabras que indican consultas estructuradas — no requieren embeddings
_STRUCTURED_KEYWORDS = [
    'region', 'año', '20', 'sector', 'tipo', 'género', 'circular',
    'sostenible', 'región', 'monto', 'financiamiento', 'cuántos',
    'cuantos', 'ranking', 'top', 'lista', 'mayor', 'menor',
]


def _is_conceptual(question: str) -> bool:
    """Heurística: retorna True si la pregunta es conceptual (no un filtro estructurado).

    Retorna False cuando la pregunta contiene palabras de filtro estructurado
    (año, región, sector, monto, rankings, etc.) que el SQL puede resolver
    directamente sin embeddings.
    """
    q_lower = question.lower()
    return not any(kw in q_lower for kw in _STRUCTURED_KEYWORDS)


# ─────────────────────────────────────────────────────────────────────────────
# CORE: llamar a Mellea con IVR
# ─────────────────────────────────────────────────────────────────────────────
def _with_backoff(fn):
    """Llama fn(). Si lanza un error 429 de rate-limit, reintenta con backoff exponencial."""
    delay = GROQ_BACKOFF_BASE
    for attempt in range(GROQ_MAX_BACKOFF_RETRIES + 1):
        try:
            return fn()
        except Exception as e:
            err_str = str(e).lower()
            is_rate_limit = '429' in err_str or 'rate_limit' in err_str or 'rate limit' in err_str
            if is_rate_limit and attempt < GROQ_MAX_BACKOFF_RETRIES:
                log.warning("Groq 429 rate-limit — esperando %ds (intento %d/%d)",
                            delay, attempt + 1, GROQ_MAX_BACKOFF_RETRIES)
                time.sleep(delay)
                delay = min(delay * 2, 60)
            else:
                raise


def _build_history_block(history: list) -> str:
    """Build a compact CONVERSATION HISTORY block from the last N turns."""
    if not history:
        return ""
    lines = ["CONVERSATION HISTORY (chronological, most recent last):"]
    for turn in reversed(history[-3:]):  # reversed so oldest first → most recent turn is closest to question
        q = turn.get("question", "")
        a = turn.get("answer", "")
        summary = turn.get("result_summary", "")
        lines.append(f"Q: {q}")
        if a:
            lines.append(f"A: {a}")
        if summary:
            lines.append(f"Data: {summary}")
        lines.append("")
    return "\n".join(lines)


def _generate_sql(question: str, history: list | None = None, semantic_question: str | None = None) -> dict:
    """
    Usa Mellea + Groq para generar SQL.
    El patrón IVR hace hasta MAX_RETRIES intentos automáticos.
    Los errores 429 de Groq se reintentan con backoff exponencial.
    Retorna un dict con sql, chart_type y opcionalmente un campo 'warning'.

    semantic_question: pregunta limpia para embeddings/keywords (usada en recovery para
    evitar que el hint de recuperación contamine la expansión semántica).
    """
    m = _get_sql_session()
    sem_q = semantic_question or question  # use clean question for semantic expansion

    # Prepend conversation history if provided
    history_block = _build_history_block(history or [])
    question_with_hint = f"{history_block}{question}" if history_block else question

    # Búsqueda semántica por embeddings (preferida) o keywords como fallback.
    # Nunca inyectar ambas: los IDs ya capturan la relevancia con más precisión que LIKE.
    semantic_ids_injected = False
    if _is_conceptual(sem_q):
        log.info("Búsqueda semántica activada para: %s", sem_q[:80])
        try:
            ids = _semantic_ids(sem_q, top_n=50)
        except Exception as e:
            log.warning("_semantic_ids falló (%s) — continuando sin embeddings.", e)
            ids = None
        if ids:
            id_list = ','.join(map(str, ids))
            question_with_hint += f"\n<!-- semantic_ids: {id_list} -->"
            log.info("Semantic IDs inyectados: %d IDs", len(ids))
            semantic_ids_injected = True
        else:
            log.info("Semantic IDs no disponibles — embeddings no cargados o tabla vacía")
    else:
        log.info("Consulta estructurada — embeddings omitidos")

    # Keywords como fallback solo cuando no hay IDs semánticos disponibles
    if not semantic_ids_injected:
        keywords = _expand_keywords(sem_q)
        if keywords:
            kw_comment = f"<!-- semantic_keywords: {', '.join(keywords)} -->"
            question_with_hint = f"{question_with_hint}\n{kw_comment}"
            log.info("Hint semántico inyectado: %s", kw_comment)

    t0 = time.perf_counter()
    result = _with_backoff(lambda: m.instruct(
        SQL_INSTRUCTION_TEMPLATE,
        user_variables={
            "question": question_with_hint,
            "relevant_fields": _retrieve_relevant_fields(question),
        },
        requirements=[
            Requirement(description="Response must be valid JSON", validation_fn=simple_validate(_is_valid_json)),
            Requirement(description="SQL must be a SELECT or null", validation_fn=simple_validate(_is_select_or_null)),
        ],
        strategy=RejectionSamplingStrategy(loop_budget=MAX_RETRIES),
        return_sampling_results=True,
    ))
    elapsed = time.perf_counter() - t0
    log.info("SQL generation: %.2fs | success=%s | retries=%d",
             elapsed,
             result.success,
             len(result.sample_generations) if result.sample_generations else 0)

    # ── extraer texto final ──────────────────────────────────────────────
    if result.success:
        raw_text = str(result.result)
    else:
        log.warning("IVR: los %d intentos no cumplieron todos los requirements. Usando fallback.", MAX_RETRIES)
        raw_text = result.sample_generations[0].value if result.sample_generations else ""

    # ── parsear JSON ─────────────────────────────────────────────────────
    try:
        parsed = json.loads(_strip_fences(raw_text))
    except (json.JSONDecodeError, ValueError):
        return {
            "sql": None,
            "chart_type": None,
            "warning": "JSON inválido tras todos los reintentos."
        }

    out = {
        "sql":        parsed.get("sql"),
        "chart_type": parsed.get("chart_type"),
    }
    if not result.success:
        out["warning"] = "Algunos requisitos no se cumplieron tras los reintentos; resultado es aproximado."
    return out


def _explain_results(question: str, sql: str, df: pd.DataFrame) -> str:
    """
    Usa Mellea + Gemma para generar una explicación en español de los resultados.
    """
    m = _get_explain_session()

    # Preparar muestra de resultados (máx 10 filas para no sobrecargar el prompt)
    sample_rows = df.head(10).to_string(index=False) if not df.empty else "(sin resultados)"
    total_rows = len(df)

    result = _with_backoff(lambda: m.instruct(
        EXPLAIN_INSTRUCTION_TEMPLATE,
        user_variables={
            "question": question,
            "sql": sql or "(ninguna)",
            "results": sample_rows,
            "total_rows": str(total_rows),
        },
    ))

    explanation = str(result).strip()
    # Limpiar posibles artefactos de formato
    explanation = _strip_fences(explanation)

    return explanation if explanation else "Los resultados se muestran en la tabla."

# ─────────────────────────────────────────────────────────────────────────────
# GUARDRAIL: validación de identificadores SQL (anti-alucinación)
# ─────────────────────────────────────────────────────────────────────────────

# Palabras clave SQL que se ignoran durante la validación de identificadores.
_SQL_KEYWORDS: set[str] = {
    "SELECT", "FROM", "WHERE", "AND", "OR", "GROUP", "BY", "ORDER", "LIMIT",
    "AS", "ON", "JOIN", "CAST", "SUM", "COUNT", "AVG", "MIN", "MAX",
    "DISTINCT", "NULL", "LIKE", "IN", "NOT", "IS", "HAVING", "INNER", "LEFT",
    "RIGHT", "OUTER", "BETWEEN", "THEN", "WHEN", "CASE", "END", "ELSE", "ASC",
    "DESC", "WITH", "UNION", "ALL", "INTEGER", "REAL", "TEXT", "TRUE", "FALSE",
    "COALESCE", "IFNULL", "IIF", "OVER", "PARTITION", "ROWID",
    # PostgreSQL type names (used in ::TYPE casts and CAST(...AS TYPE))
    "BIGINT", "FLOAT", "NUMERIC", "BOOLEAN", "TIMESTAMPTZ", "TIMESTAMP",
    "VARCHAR", "DOUBLE", "PRECISION", "INT", "SMALLINT", "DECIMAL",
    # PostgreSQL-specific functions/keywords
    "ILIKE", "SIMILAR", "TO", "EXTRACT", "EPOCH", "YEAR", "MONTH", "DAY",
    "FILTER", "WITHIN", "ARRAY", "JSON", "JSONB", "NOW", "DATE_TRUNC",
    "ROUND", "FLOOR", "CEIL", "ABS", "LENGTH", "LOWER", "UPPER", "TRIM",
    "SUBSTRING", "POSITION", "REPLACE", "NULLIF", "GREATEST", "LEAST",
    "ROW_NUMBER", "RANK", "DENSE_RANK", "LAG", "LEAD",
    "CURRENT_DATE", "CURRENT_TIMESTAMP",
}


def _get_allowed_identifiers() -> set[str]:
    """
    Construye el conjunto de identificadores válidos (tablas + columnas)
    leyendo el esquema en vivo de la base de datos.
    """
    allowed: set[str] = {"proyectos", "leads"}
    conn = get_db()
    try:
        cur = get_cursor(conn)
        for table in ("proyectos", "leads"):
            for col_name, _ in _list_table_columns(cur, table):
                allowed.add(col_name)
    finally:
        conn.close()
    return allowed


def _enforce_select_only(sql: str) -> str | None:
    """
    Verifica que el SQL sea exclusivamente SELECT o WITH (para CTEs).

    Elimina comentarios y literales de cadena antes de evaluar el primer
    keyword significativo.

    Retorna None si el SQL es seguro.
    Retorna un mensaje de error en español si detecta una sentencia de mutación
    o instrucción peligrosa (INSERT, UPDATE, DELETE, DROP, CREATE, ALTER,
    ATTACH, PRAGMA).
    """
    # Strip single-line comments (-- to end of line)
    sql_clean = re.sub(r"--[^\n]*", " ", sql)
    # Strip block comments (/* ... */)
    sql_clean = re.sub(r"/\*.*?\*/", " ", sql_clean, flags=re.DOTALL)
    # Strip string literals to prevent false negatives from embedded keywords
    sql_clean = re.sub(r"'[^']*'", " ", sql_clean)

    # Find the first meaningful keyword
    first_token = re.search(r'\b([A-Za-z]+)\b', sql_clean.strip())
    if not first_token:
        return "La consulta SQL no contiene una sentencia reconocible."

    leading_keyword = first_token.group(1).upper()

    _FORBIDDEN_KEYWORDS: set[str] = {
        "INSERT", "UPDATE", "DELETE", "DROP", "CREATE",
        "ALTER", "ATTACH", "PRAGMA",
    }

    if leading_keyword in _FORBIDDEN_KEYWORDS:
        log.warning(
            "SELECT-only guardrail: sentencia prohibida '%s' en SQL: %s",
            leading_keyword,
            sql[:120],
        )
        return (
            f"Consulta rechazada: la sentencia '{leading_keyword}' no está permitida. "
            "Solo se permiten consultas SELECT."
        )

    if leading_keyword not in ("SELECT", "WITH"):
        log.warning(
            "SELECT-only guardrail: keyword inicial inesperado '%s' en SQL: %s",
            leading_keyword,
            sql[:120],
        )
        return (
            f"Consulta rechazada: se esperaba SELECT o WITH, se encontró '{leading_keyword}'."
        )

    # Scan the full SQL for forbidden keywords to catch multi-statement bypasses
    # (e.g. "SELECT 1; DELETE FROM proyectos").
    forbidden = {'INSERT', 'UPDATE', 'DELETE', 'DROP', 'CREATE', 'ALTER', 'ATTACH', 'PRAGMA'}
    all_keywords = {m.upper() for m in re.findall(r'\b[A-Za-z]+\b', sql_clean)}
    found = forbidden & all_keywords
    if found:
        kw = next(iter(found))
        log.warning(
            "SELECT-only guardrail: sentencia prohibida '%s' encontrada en SQL completo: %s",
            kw,
            sql[:120],
        )
        return f"Consulta rechazada: sentencia '{kw}' no está permitida. Solo se aceptan consultas SELECT."

    return None


def _validate_sql_identifiers(sql: str) -> str | None:
    """
    Extrae todos los identificadores del SQL generado y los contrasta con el
    esquema real de la base de datos (tablas proyectos y leads).

    Retorna None si todos los identificadores son válidos.
    Retorna un mensaje de error en español nombrando el primer identificador
    no reconocido si alguno no está en el esquema.
    """
    allowed = _get_allowed_identifiers()

    # Extraer tokens: secuencias de caracteres alfanuméricos, guiones bajos
    # y letras con tilde (para columnas como año_adjudicacion).
    # Primero eliminar literales de cadena para evitar falsos positivos.
    sql_no_strings = re.sub(r"'[^']*'", " ", sql)

    # Recopilar todos los aliases definidos con AS (p.ej. "COUNT(*) AS cantidad").
    # Los aliases son válidos en ORDER BY / HAVING, así que se agregan al conjunto
    # de identificadores permitidos antes de validar el resto del SQL.
    defined_aliases = set(re.findall(
        r'\bAS\s+([\w\u00C0-\u017E]+)\b', sql_no_strings, flags=re.IGNORECASE
    ))
    allowed |= defined_aliases

    # Eliminar alias definidos con AS para que no sean validados dos veces.
    sql_no_strings = re.sub(r'\bAS\s+[\w\u00C0-\u017E]+\b', ' ', sql_no_strings, flags=re.IGNORECASE)

    # Extraer tokens candidatos: letras (incluidas con tilde), dígitos, guión bajo.
    raw_tokens = re.findall(r'["\`]?[\w\u00C0-\u017E]+["\`]?', sql_no_strings)

    for raw in raw_tokens:
        # Quitar comillas dobles y backticks que SQLite usa para identificadores.
        token = raw.strip('"` ')
        if not token:
            continue
        # Ignorar literales numéricos puros.
        if re.fullmatch(r'\d+(\.\d+)?', token):
            continue
        # Ignorar palabras clave SQL (comparación case-insensitive).
        if token.upper() in _SQL_KEYWORDS:
            continue
        # Ignorar alias de una sola letra (p.ej. "p" en "FROM proyectos p").
        if len(token) == 1:
            continue
        # Comprobar contra el conjunto de identificadores permitidos.
        if token not in allowed:
            log.warning("Guardrail: identificador no reconocido '%s' en SQL: %s", token, sql[:120])
            return f"Consulta rechazada: columna o identificador no reconocido '{token}'. Verifica que el campo exista en la base de datos."

    return None


# ─────────────────────────────────────────────────────────────────────────────
# AUTENTICACIÓN — decorador y rutas /api/auth/*
# ─────────────────────────────────────────────────────────────────────────────

def _get_db_user(username: str, password: str) -> dict | None:
    """Busca al usuario en la tabla `users` (unida con `organizations`).

    Retorna un dict con claves username, role, org_id y org_slug si las
    credenciales son válidas, o None si el usuario no existe o la contraseña
    es incorrecta.  Captura OperationalError para el caso en que la tabla aún
    no haya sido creada (migración pendiente).
    """
    try:
        conn = get_db()
        try:
            cur = get_cursor(conn)
            cur.execute(
                _sql(
                    """
                    SELECT u.username, u.password_hash, u.role,
                           u.org_id, o.slug
                    FROM users u
                    JOIN organizations o ON o.id = u.org_id
                    WHERE u.username = ?
                    LIMIT 1
                    """
                ),
                (username,),
            )
            row = cur.fetchone()
        finally:
            conn.close()
    except Exception as exc:
        # Captura OperationalError (tabla no existe) y cualquier otro error inesperado de DB
        log.error("_get_db_user: unexpected DB error: %s", exc)
        return None

    if row is None:
        # Ejecutar check_password_hash con un valor dummy para evitar timing attacks
        # que permitan enumerar usernames por diferencia de tiempo de respuesta
        check_password_hash(_DUMMY_HASH, password)
        return None

    # Support both sqlite3.Row (index access) and psycopg2 RealDictCursor (dict access)
    if isinstance(row, dict):
        db_username = row["username"]
        password_hash = row["password_hash"]
        role = row["role"]
        org_id = row["org_id"]
        org_slug = row["slug"]
    else:
        db_username, password_hash, role, org_id, org_slug = row
    if not check_password_hash(password_hash, password):
        return None

    return {
        "username": db_username,
        "role": role,
        "org_id": org_id,
        "org_slug": org_slug,
    }


def login_required(f):
    """Decorator que bloquea acceso a rutas protegidas si no hay sesión activa."""
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("authenticated"):
            return jsonify({"error": "No autorizado"}), 401
        return f(*args, **kwargs)
    return decorated


def role_required(role: str):
    """Decorator que exige un rol específico en la sesión (aplicar después de login_required)."""
    def decorator(f):
        @functools.wraps(f)
        def wrapped(*args, **kwargs):
            if session.get("role") != role:
                return jsonify({"error": "Acceso denegado"}), 403
            return f(*args, **kwargs)
        return wrapped
    return decorator


@app.route("/api/auth/login", methods=["POST"])
def auth_login():
    """Valida credenciales y crea sesión. Acepta JSON con username y password.

    Orden de verificación:
    1. Tabla `users` en BD (modelo org/usuario).  Si hay coincidencia, se
       almacenan org_id y org_slug en la sesión.
    2. Fallback a las credenciales en .env (comportamiento anterior).
    """
    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")

    # ── Intento 1: autenticación contra la BD ────────────────────────────────
    db_user = _get_db_user(username, password)
    if db_user is not None:
        session.clear()
        session["authenticated"] = True
        session["username"] = db_user["username"]
        session["role"] = db_user["role"]
        session["org_id"] = db_user["org_id"]
        session["org_slug"] = db_user["org_slug"]
        audit_log.info(
            "AUDIT | action=LOGIN_SUCCESS | username=%s | role=%s | org_slug=%s | source=db",
            db_user["username"], db_user["role"], db_user["org_slug"],
        )
        return jsonify({"ok": True}), 200

    # ── Intento 2: fallback a .env ───────────────────────────────────────────
    if not _ADMIN_USERNAME or not _ADMIN_PASSWORD_HASH:
        log.error("ADMIN_USERNAME o ADMIN_PASSWORD no están configurados en .env")
        return jsonify({"error": "El servidor no está configurado correctamente. Contacta al administrador."}), 500

    role = _check_credentials(username, password)
    if role is None:
        audit_log.warning("AUDIT | action=LOGIN_FAILED | username=%s", username)
        return jsonify({"error": "Usuario o contraseña incorrectos"}), 401

    session.clear()
    session["authenticated"] = True
    session["username"] = username
    session["role"] = role
    audit_log.info("AUDIT | action=LOGIN_SUCCESS | username=%s | role=%s | source=env", username, role)
    return jsonify({"ok": True}), 200


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    """Cierra la sesión actual."""
    username = session.get("username", "desconocido")
    session.clear()
    audit_log.info("AUDIT | action=LOGOUT | username=%s", username)
    return jsonify({"ok": True}), 200


@app.route("/api/auth/me", methods=["GET"])
def auth_me():
    """Retorna el estado de autenticación actual, incluyendo el rol."""
    if session.get("authenticated"):
        return jsonify({
            "authenticated": True,
            "username": session.get("username", ""),
            "role": session.get("role", "viewer"),
        }), 200
    return jsonify({"authenticated": False, "username": "", "role": None}), 200


# ─────────────────────────────────────────────────────────────────────────────
# RUTA RAÍZ — sirve el frontend automáticamente en http://localhost:5000
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/health', methods=['GET'])
def health():
    return jsonify({"status": "ok"})


@app.route('/', methods=['GET'])
def index():
    return send_file(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'corfo_app.html')
    )

# ─────────────────────────────────────────────────────────────────────────────
# CHART TYPE INFERENCE
# ─────────────────────────────────────────────────────────────────────────────

def _is_numeric_value(val) -> bool:
    """Retorna True si el valor puede interpretarse como número."""
    if val is None:
        return False
    try:
        float(val)
        return True
    except (ValueError, TypeError):
        return False


def determine_chart_type(columns: list[str], rows: list) -> str:
    """
    Inspecciona la forma del resultado SQL y devuelve el tipo de chart más adecuado.

    Reglas (en orden de prioridad):
    1. line  — exactamente 2 columnas, la primera contiene 'año', 'fecha', 'mes',
               'trimestre' o 'date'; la segunda es numérica.
    2. pie   — exactamente 2 columnas, la primera es categórica (string),
               la segunda es numérica, Y el resultado tiene ≤ 8 filas.
    3. bar   — exactamente 2 columnas, la primera es categórica,
               la segunda es numérica (cualquier cantidad de filas).
    4. table — todo lo demás (3+ columnas, columna única, segunda columna no numérica).
    """
    _TIME_KEYWORDS = ("año", "fecha", "mes", "trimestre", "date")

    if len(columns) != 2 or not rows:
        return "table"

    col0, col1 = columns[0], columns[1]

    # Tomar una muestra no-nula de la segunda columna para detectar si es numérica
    sample_val = None
    for row in rows:
        v = row[col1] if isinstance(row, dict) else (row[1] if len(row) > 1 else None)
        if v is not None:
            sample_val = v
            break

    second_is_numeric = _is_numeric_value(sample_val)

    if not second_is_numeric:
        return "table"

    # Regla 1: time series
    col0_lower = col0.lower()
    if any(kw in col0_lower for kw in _TIME_KEYWORDS):
        return "line"

    # Regla 2: pie (≤ 8 filas, categórico)
    if len(rows) <= 8:
        return "pie"

    # Regla 3: bar (categórico, cualquier tamaño)
    return "bar"


# ─────────────────────────────────────────────────────────────────────────────
# INTERPRETACIÓN HEURÍSTICA
# ─────────────────────────────────────────────────────────────────────────────
def _build_interpretacion(question: str, sql: str, num_rows: int, chart_type: str) -> str:
    """
    Construye una frase en español que explica qué entendió el sistema.
    Completamente heurística — sin LLM, sin I/O, sin red.
    Nunca lanza excepción.
    """
    _FALLBACK = "El sistema interpretó tu pregunta como una consulta sobre los proyectos CORFO."
    try:
        if not sql:
            return _FALLBACK

        sql_low = sql.lower()

        # ── Mapa de visualización ────────────────────────────────────────────
        chart_labels: dict[str, str] = {
            "bar":   "gráfico de barras",
            "line":  "gráfico de línea",
            "pie":   "gráfico circular",
            "table": "tabla",
        }
        viz = chart_labels.get(chart_type or "", "tabla")

        # ── Detectar año en cláusula WHERE ───────────────────────────────────
        year_match = re.search(r'\b(20\d{2})\b', sql_low)
        year_str = year_match.group(1) if year_match else None

        # ── Detectar tipo de agregación ──────────────────────────────────────
        has_sum   = "sum(" in sql_low
        has_count = "count(" in sql_low

        # ── Detectar GROUP BY y extraer primera columna agrupada ─────────────
        group_col: str | None = None
        gb_match = re.search(r'group\s+by\s+([\w"_\. ,]+)', sql_low)
        if gb_match:
            raw = gb_match.group(1).strip()
            # Tomar la primera columna antes de coma o salto de línea
            first = re.split(r'[,\n]', raw)[0].strip()
            # Quitar alias (tabla.columna → columna), comillas y espacios
            first = first.split('.')[-1].strip().strip('"').strip()
            if first:
                group_col = first

        # ── Componer el tema ─────────────────────────────────────────────────
        tema_parts: list[str] = []

        if has_sum and group_col:
            tema_parts.append(f"total adjudicado por {group_col}")
        elif has_count and group_col:
            tema_parts.append(f"proyectos agrupados por {group_col}")
        elif group_col:
            tema_parts.append(f"proyectos agrupados por {group_col}")
        elif has_sum:
            tema_parts.append("total adjudicado")
        elif has_count:
            tema_parts.append("conteo de proyectos")

        if year_str:
            tema_parts.append(f"del año {year_str}")

        tema = ", ".join(tema_parts) if tema_parts else "los proyectos CORFO"

        # ── Resultado en singular/plural ─────────────────────────────────────
        resultado_str = "1 resultado" if num_rows == 1 else f"{num_rows} resultado(s)"

        frase = (
            f"El sistema interpretó tu pregunta como una consulta sobre {tema}, "
            f"encontró {resultado_str} y la muestra como {viz}."
        )

        # Recortar a 200 caracteres sin cortar palabras si es necesario
        if len(frase) > 200:
            frase = frase[:197].rsplit(' ', 1)[0] + "..."

        return frase

    except Exception:
        return _FALLBACK


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES — CONSULTAS
# ─────────────────────────────────────────────────────────────────────────────
def _pg_sql(sql: str) -> str:
    """Escape literal % in AI-generated SQL for psycopg2.

    psycopg2 treats % as a parameter placeholder, so LIKE '%value%' breaks
    unless % is doubled to %%. Only needed for AI-generated SQL passed without
    params — parameterized queries use %s intentionally.
    """
    if is_postgres():
        return sql.replace("%", "%%")
    return sql


def _execute_sql_and_build_response(sql: str, question: str, chart_type: str | None, warning: str | None) -> tuple:
    """
    Ejecuta el SQL validado contra la BD y construye la respuesta JSON.

    Devuelve (response_dict, http_status_code).
    El llamador debe hacer jsonify(response_dict), status.
    """
    try:
        conn = get_db()
        df   = pd.read_sql_query(_pg_sql(sql), conn)
        conn.close()
    except Exception as e:
        log.error("SQL execution failed: %s\nSQL: %s", e, sql)
        _log_query(question, sql, None, str(e))
        return {'error': str(e)}, 500

    df = df.where(pd.notnull(df), None)

    # Descartar columna _boost si está presente (artefacto del UNION semántico)
    if '_boost' in df.columns:
        df = df.drop(columns=['_boost'])

    # Guardia de costo: truncar a 5 000 filas
    _ROW_LIMIT = 5_000
    row_limit_warning: str | None = None
    if len(df) > _ROW_LIMIT:
        log.warning("Cost guard: resultado truncado de %d a %d filas.", len(df), _ROW_LIMIT)
        df = df.iloc[:_ROW_LIMIT].copy()
        row_limit_warning = (
            "La consulta devolvió más de 5,000 filas. Se muestran solo las primeras 5,000."
        )

    warnings_parts = [w for w in (warning, row_limit_warning) if w]
    combined_warning = " | ".join(warnings_parts) if warnings_parts else None

    columns = list(df.columns)
    rows = df.to_dict(orient='records')
    inferred_chart_type = determine_chart_type(columns, rows)

    resp: dict = {
        'answer':     question,
        'dataframe':  rows,
        'columns':    columns,
        'chart_type': inferred_chart_type,
        'sql':        sql,
    }
    if combined_warning:
        resp['warning'] = combined_warning
    _log_query(question, sql, len(df), None)
    return resp, 200


@app.route('/api/query', methods=['POST'])
@login_required
def handle_query():
    ip = _get_client_ip()
    allowed, retry_after = _query_limiter.is_allowed(ip)
    if not allowed:
        log.warning("Rate limit superado: /api/query desde %s", ip)
        return jsonify({
            "error": f"Límite de solicitudes excedido. Intenta de nuevo en {retry_after} segundos."
        }), 429

    body = request.json or {}

    # ── Ruta rápida: raw_sql proporcionado directamente por el usuario ──────
    raw_sql = body.get('raw_sql', '').strip()
    if raw_sql:
        # Validar SELECT-only antes de cualquier operación con la BD
        select_only_error = _enforce_select_only(raw_sql)
        if select_only_error:
            return jsonify({'error': 'Solo se permiten consultas SELECT'}), 400

        log.info("raw_sql recibido (longitud %d chars)", len(raw_sql))
        resp, status = _execute_sql_and_build_response(
            sql=raw_sql,
            question="Query ejecutada directamente.",
            chart_type=None,
            warning=None,
        )
        if status != 200:
            return jsonify(resp), status
        # Para raw_sql no ejecutamos el explain model; usamos texto fijo
        resp['answer'] = "Query ejecutada directamente."
        return jsonify(resp)

    # ── Ruta normal: NL → SQL via LLM ────────────────────────────────────
    question = body.get('question', '').strip()
    if not question:
        return jsonify({'error': 'No question provided'}), 400

    # Historial conversacional opcional (últimas N interacciones del cliente)
    history = body.get('history', [])
    if not isinstance(history, list):
        history = []

    # 1) Pedir a Granite que genere SQL (con IVR)
    try:
        sql_out = _generate_sql(question, history=history)
    except Exception as e:
        log.exception("Error en Mellea (SQL)")
        _log_query(question, None, None, str(e))
        return jsonify({
            'answer': f'Error al comunicarse con el modelo: {str(e)}',
            'dataframe': [], 'chart_type': None, 'sql': None
        }), 500

    sql        = sql_out.get('sql')
    chart_type = sql_out.get('chart_type')
    warning    = sql_out.get('warning')

    # 2) Si no hay SQL, devolver respuesta de baja confianza
    if not sql:
        _log_query(question, None, None, "no_sql: el modelo no generó SQL")
        return jsonify(_build_fallback_response(question, "no_sql"))

    # 3a) Rechazar SQL que no sea SELECT/WITH — devuelve HTTP 403
    select_only_error = _enforce_select_only(sql)
    if select_only_error:
        return jsonify({"error": select_only_error, "sql": sql}), 403

    # 3b) Validar identificadores antes de tocar la BD (guardrail anti-alucinación)
    validation_error = _validate_sql_identifiers(sql)
    if validation_error:
        return jsonify({"error": validation_error, "sql": sql}), 422

    # 4) Ejecutar el SQL (ya validado como SELECT por Mellea y por el guardrail)
    try:
        conn = get_db()
        df   = pd.read_sql_query(_pg_sql(sql), conn)
        conn.close()
    except Exception as e:
        log.error("SQL execution failed: %s\nSQL: %s", e, sql)
        _log_query(question, sql, None, str(e))
        return jsonify({
            'answer': f'La consulta SQL generó un error al ejecutarse: {str(e)}',
            'dataframe': [], 'chart_type': None, 'sql': sql,
            'warning': 'El modelo generó SQL que no es ejecutable en esta base de datos.'
        }), 200

    df = df.where(pd.notnull(df), None)

    # 4a-bis) Guardia de costo: truncar a 5 000 filas si la BD devuelve demasiadas
    _ROW_LIMIT = 5_000
    row_limit_warning: str | None = None
    if len(df) > _ROW_LIMIT:
        log.warning("Cost guard: resultado truncado de %d a %d filas.", len(df), _ROW_LIMIT)
        df = df.iloc[:_ROW_LIMIT].copy()
        row_limit_warning = (
            "La consulta devolvió más de 5,000 filas. Se muestran solo las primeras 5,000."
        )

    # 4b) Validar calidad del resultado
    result_warning = _check_result_quality(question, df)

    # 4c) Intento de recuperación cuando el primer SQL devuelve vacío o sospechoso
    recovery_exception_logged = False
    if _is_empty_result_warning(result_warning):
        hint = (
            f"NOTA: El SQL anterior fue: [{sql}] y devolvió 0 resultados o un resultado "
            "sospechoso. Intenta una consulta más amplia: usa LIKE en vez de igualdad exacta, "
            "elimina filtros opcionales, o busca términos relacionados."
        )
        recovery_question = question + "\n\n" + hint
        sql2 = None
        try:
            sql_out2 = _generate_sql(recovery_question, semantic_question=question)
            sql2 = sql_out2.get("sql")
            chart_type2 = sql_out2.get("chart_type")

            if sql2 and sql2 != sql:
                # Validar SELECT-only y luego identificadores antes de ejecutar el SQL de recuperación
                recovery_select_error = _enforce_select_only(sql2)
                if recovery_select_error:
                    log.warning(
                        "SELECT-only guardrail rechazó SQL de recuperación (%s); se conserva resultado original.",
                        recovery_select_error,
                    )
                    sql2 = None
                recovery_validation_error = _validate_sql_identifiers(sql2) if sql2 else None
                if recovery_validation_error:
                    log.warning(
                        "Guardrail rechazó SQL de recuperación (%s); se conserva resultado original.",
                        recovery_validation_error,
                    )
                elif sql2:
                    conn2 = get_db()
                    try:
                        df2 = pd.read_sql_query(_pg_sql(sql2), conn2)
                    finally:
                        conn2.close()
                    df2 = df2.where(pd.notnull(df2), None)
                    result_warning2 = _check_result_quality(question, df2)

                    # Usar la recuperación si mejoró (sin advertencia, o al menos no vacío)
                    if not _is_empty_result_warning(result_warning2):
                        log.info(
                            "Recovery SQL produced better result (%d rows vs %d)",
                            len(df2),
                            len(df),
                        )
                        df = df2
                        sql = sql2
                        chart_type = chart_type2
                        result_warning = result_warning2
        except Exception as e:
            log.warning("Excepción durante el intento de recuperación: %s", e)
            _log_query(question, sql2, None, str(e))
            recovery_exception_logged = True
            # Se conserva el resultado original; la advertencia fluye normalmente

    # 4d) Fallback de baja confianza cuando los resultados siguen siendo vacíos/sospechosos
    if _is_empty_result_warning(result_warning):
        _ivr_signals = ("reintentos", "requisitos")
        if warning and any(s in warning for s in _ivr_signals):
            # Case B: IVR falló Y el resultado sigue vacío tras la recuperación
            _log_query(question, sql, len(df), "ivr_exhausted: resultado vacío tras reintentos")
            return jsonify(_build_fallback_response(question, "ivr_exhausted"))
        else:
            # Case C: recuperación falló sin señal IVR explícita
            _log_query(question, sql, len(df), "low_quality_result: resultado vacío o sospechoso")
            return jsonify(_build_fallback_response(question, "low_quality_result"))

    # 4e) Descartar columna _boost si está presente (artefacto del UNION semántico)
    if '_boost' in df.columns:
        df = df.drop(columns=['_boost'])

    # 5) Pedir a Gemma que explique los resultados
    try:
        answer = _explain_results(question, sql, df)
    except Exception as e:
        log.exception("Error en Mellea (Explain)")
        answer = f"Se encontraron {len(df)} resultados." if not df.empty else "No se encontraron resultados."

    # Combinar advertencia del IVR, la de calidad del resultado y la del límite de filas
    warnings_parts = [w for w in (warning, result_warning, row_limit_warning) if w]
    combined_warning = " | ".join(warnings_parts) if warnings_parts else None

    columns = list(df.columns)
    rows = df.to_dict(orient='records')
    inferred_chart_type = determine_chart_type(columns, rows)

    resp = {
        'answer':         answer,
        'dataframe':      rows,
        'columns':        columns,
        'chart_type':     inferred_chart_type,
        'sql':            sql,
        'interpretacion': _build_interpretacion(question, sql, len(rows), inferred_chart_type),
    }
    if combined_warning:
        resp['warning'] = combined_warning
    # Indicar baja confianza cuando el IVR no pudo satisfacer todos los requisitos
    _ivr_signals = ("reintentos", "requisitos")
    if warning and any(s in warning for s in _ivr_signals):
        resp['low_confidence'] = True
    if not recovery_exception_logged:
        _log_query(question, sql, len(df), None)
    return jsonify(resp)

# ─────────────────────────────────────────────────────────────────────────────
# ROUTES — EXCEL EXPORT  (genérico: cualquier dataframe)
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/export/excel', methods=['POST'])
@login_required
@role_required("admin")
def export_excel():
    body     = request.json or {}
    rows     = body.get('data', [])
    filename = body.get('filename', 'exportacion_corfo')

    if not rows:
        return jsonify({'error': 'No data'}), 400

    df   = pd.DataFrame(rows)
    path = f'/tmp/{filename}.xlsx'

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Datos')
        ws = writer.sheets['Datos']
        for i, col in enumerate(df.columns, 1):
            max_len = max(df[col].astype(str).str.len().max(), len(str(col)))
            # openpyxl usa letras A-Z; para más de 26 columnas usar get_column_letter
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 45)

    audit_log.info("AUDIT | action=EXPORT_EXCEL | ip=%s | rows=%d", _get_client_ip(), len(rows))
    return send_file(path, as_attachment=True, download_name=f'{filename}.xlsx')

# ─────────────────────────────────────────────────────────────────────────────
# ROUTES — CSV EXPORT  (genérico: cualquier dataframe)
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/export/csv', methods=['POST'])
@login_required
@role_required("admin")
def export_csv():
    body     = request.json or {}
    rows     = body.get('data', [])
    filename = body.get('filename', 'exportacion_corfo')

    if not rows:
        return jsonify({'error': 'No data'}), 400

    df      = pd.DataFrame(rows)
    csv_str = df.to_csv(index=False)
    # UTF-8 BOM so Excel on Windows opens the file without encoding issues
    csv_bytes = ('\ufeff' + csv_str).encode('utf-8')
    buf = io.BytesIO(csv_bytes)
    buf.seek(0)

    audit_log.info("AUDIT | action=EXPORT_CSV | ip=%s | rows=%d", _get_client_ip(), len(rows))
    response = Response(
        buf.read(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}.csv"'}
    )
    return response

# ─────────────────────────────────────────────────────────────────────────────
# ROUTES — DASHBOARD (datos pre-agrupados para los 8 gráficos)
# ─────────────────────────────────────────────────────────────────────────────

def _build_dashboard_where(args: dict) -> tuple[str, list]:
    """Construye cláusula WHERE y lista de parámetros a partir de query params del dashboard.
    Retorna (where_clause, params_list). where_clause incluye 'WHERE' si aplica, o '' si no hay filtros.
    """
    conditions: list[str] = []
    params: list = []

    año_desde = args.get('año_desde', '').strip()
    año_hasta = args.get('año_hasta', '').strip()
    region    = args.get('region', '').strip()
    sector    = args.get('sector', '').strip()
    tipo_inn  = args.get('tipo_innovacion', '').strip()
    tipo_proy = args.get('tipo_proyecto', '').strip()

    if año_desde:
        try:
            conditions.append(_sql('"año_adjudicacion" >= ?'))
            params.append(int(año_desde))
        except ValueError:
            pass
    if año_hasta:
        try:
            conditions.append(_sql('"año_adjudicacion" <= ?'))
            params.append(int(año_hasta))
        except ValueError:
            pass
    if region:
        conditions.append(_sql('region_ejecucion = ?'))
        params.append(region)
    if sector:
        conditions.append(_sql('sector_economico = ?'))
        params.append(sector)
    if tipo_inn:
        conditions.append(_sql('tipo_innovacion = ?'))
        params.append(tipo_inn)
    if tipo_proy:
        conditions.append(_sql('tipo_proyecto = ?'))
        params.append(tipo_proy)

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''
    return where, params


@app.route('/api/dashboard/filter-options', methods=['GET'])
@login_required
def dashboard_filter_options():
    """Retorna los valores distintos de cada campo filtrable para poblar los dropdowns."""
    conn = get_db()
    try:
        def distinct(sql: str, col: str) -> list:
            df = pd.read_sql_query(sql, conn)
            return [v for v in df[col].tolist() if v is not None and not (isinstance(v, float) and math.isnan(v))]

        regiones = distinct(
            "SELECT DISTINCT region_ejecucion FROM proyectos WHERE region_ejecucion IS NOT NULL ORDER BY region_ejecucion",
            'region_ejecucion'
        )
        sectores = distinct(
            "SELECT DISTINCT sector_economico FROM proyectos WHERE sector_economico IS NOT NULL ORDER BY sector_economico",
            'sector_economico'
        )
        tipos_innovacion = distinct(
            "SELECT DISTINCT tipo_innovacion FROM proyectos WHERE tipo_innovacion IS NOT NULL ORDER BY tipo_innovacion",
            'tipo_innovacion'
        )
        tipos_proyecto = distinct(
            "SELECT DISTINCT tipo_proyecto FROM proyectos WHERE tipo_proyecto IS NOT NULL ORDER BY tipo_proyecto",
            'tipo_proyecto'
        )
        años = distinct(
            'SELECT DISTINCT "año_adjudicacion" FROM proyectos WHERE "año_adjudicacion" IS NOT NULL ORDER BY "año_adjudicacion"',
            'año_adjudicacion'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

    return jsonify({
        'regiones': regiones,
        'sectores': sectores,
        'tipos_innovacion': tipos_innovacion,
        'tipos_proyecto': tipos_proyecto,
        'año_min': int(min(años)) if años else 2009,
        'año_max': int(max(años)) if años else 2025,
    })


@app.route('/api/dashboard', methods=['GET'])
@login_required
def dashboard_data():
    # Validate year range before hitting the DB
    año_desde_raw = request.args.get('año_desde', '').strip()
    año_hasta_raw = request.args.get('año_hasta', '').strip()
    if año_desde_raw and año_hasta_raw:
        try:
            if int(año_desde_raw) > int(año_hasta_raw):
                return jsonify({'error': 'año_desde no puede ser mayor que año_hasta'}), 400
        except ValueError:
            pass

    conn = get_db()
    try:
        where, params = _build_dashboard_where(request.args)

        def q(sql: str, extra_params: list | None = None) -> list:
            all_params = params + (extra_params or [])
            df = pd.read_sql_query(sql, conn, params=all_params if all_params else None)
            records = df.to_dict('records')
            # Replace float NaN/inf with None so Flask produces valid JSON
            return [
                {k: (None if isinstance(v, float) and (math.isnan(v) or math.isinf(v)) else v)
                 for k, v in row.items()}
                for row in records
            ]

        by_year = q(f"""
            SELECT "año_adjudicacion" as año,
                   COUNT(*) as proyectos,
                   CAST(SUM(CAST(aprobado_corfo AS REAL)) AS BIGINT) as monto_total
            FROM proyectos {where} GROUP BY "año_adjudicacion" ORDER BY "año_adjudicacion"
        """)
        by_region = q(f"""
            SELECT region_ejecucion as region,
                   COUNT(*) as proyectos,
                   CAST(SUM(CAST(aprobado_corfo AS REAL)) AS BIGINT) as monto_total
            FROM proyectos {where} GROUP BY region_ejecucion
            ORDER BY proyectos DESC LIMIT 10
        """)
        by_sector = q(f"""
            SELECT sector_economico as sector,
                   COUNT(*) as proyectos
            FROM proyectos {where} GROUP BY sector_economico
            ORDER BY proyectos DESC LIMIT 8
        """)

        # genero_director: combinar el filtro base con la condición NOT NULL
        gender_where = (where + ' AND genero_director IS NOT NULL') if where else 'WHERE genero_director IS NOT NULL'
        by_gender = q(f"""
            SELECT genero_director as genero,
                   COUNT(*) as proyectos
            FROM proyectos {gender_where}
            GROUP BY genero_director
        """)

        circular = q(f"""
            SELECT economia_circular_si_no as circular,
                   COUNT(*) as proyectos
            FROM proyectos {where} GROUP BY economia_circular_si_no
        """)

        # tipo_innovacion: combinar el filtro base con la condición NOT NULL
        innov_where = (where + ' AND tipo_innovacion IS NOT NULL') if where else 'WHERE tipo_innovacion IS NOT NULL'
        by_innovation = q(f"""
            SELECT tipo_innovacion as tipo,
                   COUNT(*) as proyectos,
                   CAST(SUM(CAST(aprobado_corfo AS REAL)) AS BIGINT) as monto_total
            FROM proyectos {innov_where}
            GROUP BY tipo_innovacion ORDER BY monto_total DESC
        """)

        # top companies: combinar el filtro base con la condición NOT NULL / != Persona Natural
        companies_where = (where + " AND razon IS NOT NULL AND razon != 'Persona Natural'") if where else "WHERE razon IS NOT NULL AND razon != 'Persona Natural'"
        top_companies = q(f"""
            SELECT razon as empresa,
                   COUNT(*) as proyectos,
                   CAST(SUM(CAST(aprobado_corfo AS REAL)) AS BIGINT) as monto_total
            FROM proyectos {companies_where}
            GROUP BY razon ORDER BY monto_total DESC LIMIT 10
        """)

        # tramo_ventas: combinar el filtro base con la condición NOT NULL
        tramo_where = (where + ' AND tramo_ventas IS NOT NULL') if where else 'WHERE tramo_ventas IS NOT NULL'
        by_tramo = q(f"""
            SELECT tramo_ventas as tramo,
                   COUNT(*) as proyectos
            FROM proyectos {tramo_where}
            GROUP BY tramo_ventas ORDER BY proyectos DESC
        """)

        return jsonify({
            'by_year': by_year, 'by_region': by_region,
            'by_sector': by_sector, 'by_gender': by_gender,
            'circular': circular, 'by_innovation': by_innovation,
            'top_companies': top_companies, 'by_tramo': by_tramo
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# ─────────────────────────────────────────────────────────────────────────────
# ROUTES — LEADS (CRM)
# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# ROUTES — DASHBOARD DRILL-DOWN (filas individuales de proyectos para un valor)
# ─────────────────────────────────────────────────────────────────────────────
_DRILL_ALLOWED_FIELDS: set[str] = {
    'region_ejecucion', 'sector_economico', 'tipo_innovacion', 'tipo_proyecto',
    'tendencia_final', 'sostenible', 'economia_circular_si_no', 'año_adjudicacion',
    'genero_director', 'tramo_ventas', 'razon',
}

@app.route('/api/dashboard/drill', methods=['GET'])
@login_required
def dashboard_drill():
    field = request.args.get('field', '').strip()
    value = request.args.get('value', '').strip()

    if not field:
        return jsonify({'error': 'El parámetro field es requerido'}), 400
    if not value:
        return jsonify({'error': 'El parámetro value es requerido'}), 400
    if field not in _DRILL_ALLOWED_FIELDS:
        return jsonify({'error': f"Campo no permitido: '{field}'. Campos válidos: {', '.join(sorted(_DRILL_ALLOWED_FIELDS))}"}), 400

    conn = get_db()
    try:
        # año_adjudicacion must be double-quoted; it is INTEGER so compare directly
        if field == 'año_adjudicacion':
            sql = _sql(
                'SELECT razon, aprobado_corfo, "año_adjudicacion", '
                'region_ejecucion, sector_economico, tipo_proyecto '
                'FROM proyectos '
                'WHERE "año_adjudicacion" = ? '
                'ORDER BY "año_adjudicacion" DESC '
                'LIMIT 200'
            )
            try:
                param: object = int(value)
            except (ValueError, TypeError):
                conn.close()
                return jsonify({'error': 'El valor para año_adjudicacion debe ser un número entero'}), 400
        else:
            sql = _sql(
                f'SELECT razon, aprobado_corfo, "año_adjudicacion", '
                f'region_ejecucion, sector_economico, tipo_proyecto '
                f'FROM proyectos '
                f'WHERE "{field}" = ? '
                f'ORDER BY "año_adjudicacion" DESC '
                f'LIMIT 200'
            )
            param = value

        df = pd.read_sql_query(sql, conn, params=[param])
        total = int(df.shape[0])
        rows = df.where(pd.notnull(df), None).to_dict('records')
    except Exception as e:
        log.error("Error en dashboard/drill field=%s value=%s: %s", field, value, e)
        conn.close()
        return jsonify({'error': f'Error al consultar la base de datos: {str(e)}'}), 500

    conn.close()
    return jsonify({'rows': rows, 'field': field, 'value': value, 'total': total})


@app.route('/api/empresa/<path:razon>', methods=['GET'])
@login_required
def empresa_perfil(razon: str):
    """Devuelve el historial completo de proyectos de una empresa por razón social."""
    audit_log.info("EMPRESA_PERFIL | razon=%s | ip=%s", razon, _get_client_ip())

    conn = get_db()
    try:
        df = pd.read_sql_query(
            _sql(
                """
                SELECT codigo,
                       "año_adjudicacion",
                       instrumento,
                       titulo_del_proyecto,
                       CAST(aprobado_corfo AS REAL) AS aprobado_corfo,
                       sector_economico,
                       region_ejecucion,
                       tendencia_final
                FROM proyectos
                WHERE razon = ?
                ORDER BY "año_adjudicacion" DESC
                """
            ),
            conn,
            params=[razon],
        )
    except Exception as e:
        conn.close()
        log.error("Error en empresa_perfil razon=%s: %s", razon, e)
        return jsonify({'error': f'Error al consultar la base de datos: {str(e)}'}), 500

    conn.close()

    if df.empty:
        return jsonify({'error': 'Empresa no encontrada'}), 404

    def _clean(v):
        """Reemplaza float NaN/inf con None para que jsonify produzca JSON válido."""
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v

    proyectos = [
        {k: _clean(row[k]) for k in df.columns}
        for row in df.to_dict('records')
    ]

    def _to_float(v):
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    # Estadísticas globales
    montos = [_to_float(r['aprobado_corfo']) for r in proyectos if r['aprobado_corfo'] is not None]
    montos = [m for m in montos if m is not None]
    monto_total = int(sum(montos)) if montos else 0

    años = [r['año_adjudicacion'] for r in proyectos if r['año_adjudicacion'] is not None]
    primer_año = int(min(años)) if años else None
    ultimo_año  = int(max(años)) if años else None

    # Agregado por año para el gráfico de línea
    por_año_map: dict[int, dict] = {}
    for r in proyectos:
        año = r['año_adjudicacion']
        if año is None:
            continue
        año = int(año)
        if año not in por_año_map:
            por_año_map[año] = {'año': año, 'proyectos': 0, 'monto': 0}
        por_año_map[año]['proyectos'] += 1
        monto_r = _to_float(r['aprobado_corfo'])
        if monto_r is not None:
            por_año_map[año]['monto'] += int(monto_r)
    por_año = sorted(por_año_map.values(), key=lambda x: x['año'])

    return jsonify({
        'razon':           razon,
        'total_proyectos': len(proyectos),
        'monto_total':     monto_total,
        'primer_año':      primer_año,
        'ultimo_año':      ultimo_año,
        'proyectos':       proyectos,
        'por_año':         por_año,
    })


LEADS_STATUSES = ['Nuevo', 'Contactado', 'En seguimiento', 'Propuesta enviada', 'Cerrado']


def _migrate_leads_table() -> None:
    """Migra la tabla leads al nuevo schema simplificado si detecta el schema antiguo.

    Only runs on SQLite — PostgreSQL schema migrations are managed separately
    via schema_migrations/ (DOB-188).
    """
    if is_postgres():
        log.info("_migrate_leads_table: omitido en modo PostgreSQL (gestionar via migraciones).")
        return

    conn = sqlite3.connect(os.environ.get("DB_PATH", DB), timeout=10)
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(leads)")
    cols = {row[1] for row in cur.fetchall()}

    if 'rut_beneficiario' in cols:
        log.info("Detectado schema antiguo de leads — renombrando a leads_old…")
        cur.execute("DROP TABLE IF EXISTS leads_old")
        cur.execute("ALTER TABLE leads RENAME TO leads_old")
        log.info("leads_old creado como backup del schema anterior.")
        cols = set()  # tabla renombrada, crear la nueva

    if 'nombre_compania' not in cols:
        cur.execute(
            "CREATE TABLE IF NOT EXISTS leads ("
            "  id               INTEGER PRIMARY KEY AUTOINCREMENT,"
            "  user_id          TEXT    NOT NULL DEFAULT 'default',"
            "  nombre_compania  TEXT    NOT NULL,"
            "  contacto         TEXT,"
            "  status           TEXT    NOT NULL DEFAULT 'Nuevo',"
            "  notas            TEXT,"
            "  created_at       TEXT    DEFAULT (datetime('now')),"
            "  updated_at       TEXT    DEFAULT (datetime('now'))"
            ")"
        )
        cur.execute("CREATE INDEX IF NOT EXISTS idx_leads_user ON leads(user_id)")
        conn.commit()
        log.info("Nueva tabla leads creada con schema simplificado.")
    else:
        # Migración defensiva: agregar columna notas si no existe
        if 'notas' not in cols:
            try:
                cur.execute("ALTER TABLE leads ADD COLUMN notas TEXT DEFAULT NULL")
                conn.commit()
                log.info("Columna 'notas' agregada a tabla leads.")
            except Exception as e:
                log.warning("No se pudo agregar columna 'notas' a leads: %s", e)

    conn.close()


def _ensure_actividad_table() -> None:
    """Crea la tabla actividad si no existe, en SQLite o PostgreSQL.
    También agrega la columna notas a leads en PostgreSQL si no existe.
    """
    conn = get_db()
    try:
        cur = get_cursor(conn)

        # Migración defensiva: agregar columna notas a leads en PostgreSQL
        if is_postgres():
            try:
                cur.execute(
                    "ALTER TABLE leads ADD COLUMN IF NOT EXISTS notas TEXT DEFAULT NULL"
                )
                conn.commit()
                log.info("Columna 'notas' verificada/agregada a tabla leads (PostgreSQL).")
            except Exception as e:
                conn.rollback()
                log.warning("No se pudo agregar columna 'notas' a leads (PostgreSQL): %s", e)

        if is_postgres():
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS actividad (
                    id              SERIAL PRIMARY KEY,
                    nombre_compania TEXT NOT NULL,
                    fecha           TEXT NOT NULL,
                    tipo            TEXT NOT NULL,
                    con_quien       TEXT,
                    nota            TEXT,
                    user_id         TEXT NOT NULL DEFAULT 'default',
                    created_at      TEXT DEFAULT (to_char(now(), 'YYYY-MM-DD"T"HH24:MI:SS'))
                )
                """
            )
        else:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS actividad (
                    id              INTEGER PRIMARY KEY AUTOINCREMENT,
                    nombre_compania TEXT NOT NULL,
                    fecha           TEXT NOT NULL,
                    tipo            TEXT NOT NULL,
                    con_quien       TEXT,
                    nota            TEXT,
                    user_id         TEXT NOT NULL DEFAULT 'default',
                    created_at      TEXT DEFAULT (datetime('now'))
                )
                """
            )
        conn.commit()
        log.info("Tabla actividad verificada/creada OK.")
    except Exception as e:
        log.error("Error al crear tabla actividad: %s", e)
        conn.rollback()
    finally:
        conn.close()


@app.route('/api/leads', methods=['GET'])
@login_required
def get_leads():
    user_id = session.get('username', 'default')
    conn = get_db()
    df = pd.read_sql_query(
        _sql("SELECT * FROM leads WHERE user_id = ? ORDER BY updated_at DESC LIMIT 500"),
        conn, params=[user_id]
    )
    conn.close()
    return jsonify(df.where(pd.notnull(df), None).to_dict('records'))


@app.route('/api/leads/stats', methods=['GET'])
@login_required
def get_leads_stats():
    user_id = session.get('username', 'default')
    conn = get_db()
    c = get_cursor(conn)
    c.execute(_sql("SELECT COUNT(*) FROM leads WHERE user_id = ?"), [user_id])
    row = c.fetchone()
    total = row[0] if not isinstance(row, dict) else list(row.values())[0]
    c.execute(_sql("SELECT status, COUNT(*) AS count FROM leads WHERE user_id = ? GROUP BY status"), [user_id])
    if is_postgres():
        by_status = {r["status"]: r["count"] for r in c.fetchall()}
    else:
        by_status = dict(c.fetchall())
    conn.close()
    contacted = sum(v for k, v in by_status.items() if k != 'Nuevo')
    return jsonify({
        'total': total,
        'by_status': by_status,
        'contacted': contacted,
        'pending': by_status.get('Nuevo', 0),
    })


@app.route('/api/leads/<int:lid>', methods=['GET'])
@login_required
def get_lead(lid):
    user_id = session.get('username', 'default')
    conn = get_db()
    df = pd.read_sql_query(
        _sql("SELECT * FROM leads WHERE id = ? AND user_id = ?"),
        conn, params=[lid, user_id]
    )
    conn.close()
    if df.empty:
        return jsonify({'error': 'Lead no encontrado'}), 404
    return jsonify(df.where(pd.notnull(df), None).to_dict('records')[0])


@app.route('/api/leads', methods=['POST'])
@login_required
@role_required("admin")
def create_lead():
    data = request.json or {}
    # Accept 'nombre_compania' or legacy 'razon' from the Consultas tab
    nombre = (data.get('nombre_compania') or data.get('razon') or '').strip()
    if not nombre:
        return jsonify({'error': 'nombre_compania requerido'}), 400
    if nombre == 'Persona Natural':
        return jsonify({'error': 'Persona Natural no puede agregarse como lead'}), 400

    user_id = session.get('username', 'default')
    contacto = (data.get('contacto') or '').strip()
    status = data.get('status', 'Nuevo')
    if status not in LEADS_STATUSES:
        status = 'Nuevo'

    conn = get_db()
    existing = pd.read_sql_query(
        _sql("SELECT id FROM leads WHERE user_id = ? AND nombre_compania = ?"),
        conn, params=[user_id, nombre]
    )
    if not existing.empty:
        conn.close()
        return jsonify({'error': 'La empresa ya está en leads', 'id': int(existing.iloc[0]['id'])}), 409

    c = get_cursor(conn)
    try:
        if is_postgres():
            c.execute(
                _sql(
                    "INSERT INTO leads (user_id, nombre_compania, contacto, status) "
                    "VALUES (?, ?, ?, ?) RETURNING id"
                ),
                [user_id, nombre, contacto, status],
            )
            row = c.fetchone()
            new_id = row["id"] if isinstance(row, dict) else row[0]
        else:
            c.execute(
                "INSERT INTO leads (user_id, nombre_compania, contacto, status) VALUES (?, ?, ?, ?)",
                [user_id, nombre, contacto, status],
            )
            new_id = c.lastrowid
        conn.commit()
    except Exception as e:
        conn.close()
        log.error("Error al crear lead para %s: %s", nombre, e)
        return jsonify({'error': str(e)}), 500

    conn.close()
    audit_log.info("AUDIT | action=LEAD_CREATED | ip=%s | user=%s | empresa=%s",
                   _get_client_ip(), user_id, nombre)
    return jsonify({'success': True, 'id': new_id, 'nombre_compania': nombre}), 201


@app.route('/api/leads/<int:lid>', methods=['PUT'])
@login_required
@role_required("admin")
def update_lead(lid):
    user_id = session.get('username', 'default')
    data = request.json or {}
    sets, params = [], []
    ph = "%s" if is_postgres() else "?"
    for field in ('contacto', 'status', 'notas'):
        if field not in data:
            continue
        if field == 'status' and data[field] not in LEADS_STATUSES:
            continue
        sets.append(f"{field} = {ph}")
        params.append(data[field])
    if not sets:
        # No valid fields supplied — still verify the lead exists
        conn = get_db()
        c = get_cursor(conn)
        c.execute(_sql("SELECT 1 FROM leads WHERE id = ? AND user_id = ?"), [lid, user_id])
        exists = c.fetchone()
        conn.close()
        if not exists:
            return jsonify({'error': 'Lead no encontrado'}), 404
        return jsonify({'success': True})
    sets.append(f"updated_at = {ph}")
    params.append(datetime.now().isoformat())
    params.extend([lid, user_id])
    conn = get_db()
    c = get_cursor(conn)
    c.execute(f"UPDATE leads SET {', '.join(sets)} WHERE id = {ph} AND user_id = {ph}", params)
    conn.commit()
    if c.rowcount == 0:
        conn.close()
        return jsonify({'error': 'Lead no encontrado'}), 404
    conn.close()
    audit_log.info("AUDIT | action=LEAD_UPDATED | ip=%s | user=%s | lead_id=%d",
                   _get_client_ip(), user_id, lid)
    return jsonify({'success': True})


@app.route('/api/leads/<int:lid>', methods=['DELETE'])
@login_required
@role_required("admin")
def delete_lead(lid):
    user_id = session.get('username', 'default')
    conn = get_db()
    c = get_cursor(conn)
    c.execute(_sql("DELETE FROM leads WHERE id = ? AND user_id = ?"), [lid, user_id])
    conn.commit()
    deleted = c.rowcount
    conn.close()
    if deleted == 0:
        return jsonify({'error': 'Lead no encontrado'}), 404
    audit_log.info("AUDIT | action=LEAD_DELETED | ip=%s | user=%s | lead_id=%d",
                   _get_client_ip(), user_id, lid)
    return jsonify({'success': True})


@app.route('/api/leads/export/csv', methods=['GET'])
@login_required
@role_required("admin")
def export_leads_csv():
    user_id = session.get('username', 'default')
    conn = get_db()
    df = pd.read_sql_query(
        _sql(
            "SELECT nombre_compania, contacto, status, created_at, updated_at "
            "FROM leads WHERE user_id = ? ORDER BY updated_at DESC"
        ),
        conn, params=[user_id]
    )
    conn.close()
    output = io.StringIO()
    df.to_csv(output, index=False, encoding='utf-8')
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=leads_corfo.csv'}
    )

# ─────────────────────────────────────────────────────────────────────────────
# ROUTES — EXPLORADOR DE PROYECTOS (tabla paginada con filtros dinámicos)
# ─────────────────────────────────────────────────────────────────────────────

# Columnas permitidas como sort_col (whitelist para evitar inyección SQL)
_EXPLORADOR_SORT_WHITELIST: set[str] = {
    'año_adjudicacion', 'razon', 'titulo_del_proyecto', 'region_ejecucion',
    'sector_economico', 'tipo_innovacion', 'tipo_proyecto', 'aprobado_corfo',
    'tendencia_final', 'estado_data', 'instrumento_homologado', 'tramo_ventas',
    'codigo',
}


@app.route('/api/proyectos', methods=['GET'])
@login_required
def explorador_proyectos():
    """Endpoint paginado para el Explorador de Proyectos.

    Soporta filtros por región, sector, tipo_innovacion, tipo_proyecto,
    rango de años, rango de monto CORFO, búsqueda libre (razon + titulo),
    ordenamiento y paginación.
    """
    args = request.args

    # ── Paginación ──────────────────────────────────────────────────────────
    try:
        page = max(1, int(args.get('page', 1)))
    except (ValueError, TypeError):
        page = 1
    try:
        limit = min(200, max(1, int(args.get('limit', 50))))
    except (ValueError, TypeError):
        limit = 50
    offset = (page - 1) * limit

    # ── Ordenamiento (usar whitelist para sort_col) ──────────────────────────
    sort_col = args.get('sort_col', 'año_adjudicacion').strip()
    if sort_col not in _EXPLORADOR_SORT_WHITELIST:
        sort_col = 'año_adjudicacion'
    sort_dir = 'DESC' if args.get('sort_dir', 'desc').lower() == 'desc' else 'ASC'

    # ── Construcción dinámica del WHERE ─────────────────────────────────────
    conditions: list[str] = []
    params: list = []

    region     = args.get('region', '').strip()
    sector     = args.get('sector', '').strip()
    tipo_inn   = args.get('tipo_innovacion', '').strip()
    tipo_proy  = args.get('tipo_proyecto', '').strip()
    search     = args.get('search', '').strip()

    if region:
        conditions.append(_sql('region_ejecucion = ?'))
        params.append(region)
    if sector:
        conditions.append(_sql('sector_economico = ?'))
        params.append(sector)
    if tipo_inn:
        conditions.append(_sql('tipo_innovacion = ?'))
        params.append(tipo_inn)
    if tipo_proy:
        conditions.append(_sql('tipo_proyecto = ?'))
        params.append(tipo_proy)

    # Rango de años (año_adjudicacion siempre entre comillas dobles)
    anio_min = args.get('anio_min', '').strip()
    anio_max = args.get('anio_max', '').strip()
    if anio_min:
        try:
            conditions.append(_sql('"año_adjudicacion" >= ?'))
            params.append(int(anio_min))
        except ValueError:
            pass
    if anio_max:
        try:
            conditions.append(_sql('"año_adjudicacion" <= ?'))
            params.append(int(anio_max))
        except ValueError:
            pass

    # Rango de monto CORFO — TEXT en SQLite, NUMERIC en PostgreSQL
    monto_min = args.get('monto_min', '').strip()
    monto_max = args.get('monto_max', '').strip()
    if monto_min:
        try:
            v = float(monto_min)
            if is_postgres():
                conditions.append(_sql('aprobado_corfo >= ?'))
            else:
                conditions.append(_sql('CAST(aprobado_corfo AS REAL) >= ?'))
            params.append(v)
        except ValueError:
            pass
    if monto_max:
        try:
            v = float(monto_max)
            if is_postgres():
                conditions.append(_sql('aprobado_corfo <= ?'))
            else:
                conditions.append(_sql('CAST(aprobado_corfo AS REAL) <= ?'))
            params.append(v)
        except ValueError:
            pass

    # Búsqueda libre en razon Y titulo_del_proyecto
    if search:
        like_val = f'%{search}%'
        conditions.append(_sql('(razon LIKE ? OR titulo_del_proyecto LIKE ?)'))
        params.extend([like_val, like_val])

    where_clause = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''

    # ── Expresión de ORDER BY (sort_col validado contra whitelist) ───────────
    # aprobado_corfo requiere CAST en SQLite para orden numérico correcto
    if sort_col == 'aprobado_corfo' and not is_postgres():
        order_expr = f'CAST(aprobado_corfo AS REAL) {sort_dir}'
    elif sort_col == 'año_adjudicacion':
        order_expr = f'"año_adjudicacion" {sort_dir}'
    else:
        order_expr = f'"{sort_col}" {sort_dir}'

    conn = get_db()
    try:
        # COUNT total para calcular páginas
        count_sql = f'SELECT COUNT(*) AS total FROM proyectos {where_clause}'
        count_df = pd.read_sql_query(count_sql, conn, params=params if params else None)
        total = int(count_df.iloc[0]['total'])

        # Fetch de datos paginados
        data_sql = (
            f'SELECT codigo, "año_adjudicacion", razon, titulo_del_proyecto, '
            f'region_ejecucion, sector_economico, tipo_innovacion, tipo_proyecto, '
            f'{"aprobado_corfo" if is_postgres() else "CAST(aprobado_corfo AS REAL)"} AS aprobado_corfo, '
            f'tendencia_final, estado_data, instrumento_homologado, tramo_ventas '
            f'FROM proyectos {where_clause} '
            f'ORDER BY {order_expr} '
            f'LIMIT {limit} OFFSET {offset}'
        )
        df = pd.read_sql_query(data_sql, conn, params=params if params else None)
        # Use pandas JSON serialization to safely convert numpy/Decimal types
        import json as _json
        rows = _json.loads(df.to_json(orient='records', force_ascii=False, default_handler=str))

        pages = max(1, math.ceil(total / limit))
        return jsonify({
            'rows':  rows,
            'total': total,
            'page':  page,
            'limit': limit,
            'pages': pages,
        })
    except Exception as e:
        log.error("Error en /api/proyectos: %s", e, exc_info=True)
        return jsonify({'error': f'Error al consultar la base de datos: {str(e)}'}), 500
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES — PROYECTO DETALLE
# ─────────────────────────────────────────────────────────────────────────────

def _proyecto_row_to_dict(row) -> dict:
    """Convierte una fila de proyectos a dict, casteando aprobado_corfo a float."""
    d = dict(row)
    for money_col in ('aprobado_corfo', 'aprobado_privado', 'aprobado_privado_pecuniario', 'monto_consolidado_ley'):
        if money_col in d and d[money_col] is not None:
            try:
                d[money_col] = float(d[money_col])
            except (ValueError, TypeError):
                d[money_col] = None
    # Replace float NaN/inf with None for safe JSON serialization
    for k, v in d.items():
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            d[k] = None
    return d


@app.route('/api/proyecto/by-title', methods=['GET'])
@login_required
def proyecto_by_title():
    """Busca proyectos por título parcial (LIKE). Devuelve hasta 5 coincidencias."""
    titulo = request.args.get('titulo', '').strip()
    if not titulo:
        return jsonify({'error': 'Parámetro titulo requerido'}), 400

    conn = get_db()
    try:
        c = get_cursor(conn)
        c.execute(
            _sql(
                """
                SELECT *,
                       CAST(aprobado_corfo AS REAL) AS aprobado_corfo
                FROM proyectos
                WHERE titulo_del_proyecto LIKE ?
                LIMIT 5
                """
            ),
            [f'%{titulo}%'],
        )
        rows = c.fetchall()
    except Exception as e:
        log.error("Error en proyecto_by_title titulo=%s: %s", titulo, e)
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

    return jsonify([_proyecto_row_to_dict(r) for r in rows])


@app.route('/api/proyecto/<path:codigo>', methods=['GET'])
@login_required
def proyecto_detalle(codigo: str):
    """Devuelve todos los campos de un proyecto por su código."""
    audit_log.info("PROYECTO_DETALLE | codigo=%s | ip=%s", codigo, _get_client_ip())

    conn = get_db()
    try:
        c = get_cursor(conn)
        c.execute(
            _sql(
                """
                SELECT *,
                       CAST(aprobado_corfo AS REAL) AS aprobado_corfo
                FROM proyectos
                WHERE codigo = ?
                LIMIT 1
                """
            ),
            [codigo],
        )
        row = c.fetchone()
    except Exception as e:
        log.error("Error en proyecto_detalle codigo=%s: %s", codigo, e)
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

    if row is None:
        return jsonify({'error': 'Proyecto no encontrado'}), 404

    return jsonify(_proyecto_row_to_dict(row))


# ─────────────────────────────────────────────────────────────────────────────
# CRM OBJECT MODEL
# ─────────────────────────────────────────────────────────────────────────────

def _slugify(text: str) -> str:
    """Convierte una razón social en un slug apto para crm_id.

    Lowercase, reemplaza espacios y caracteres no alfanuméricos con guiones,
    colapsa guiones consecutivos y elimina guiones al inicio/fin.
    """
    slug = text.lower()
    slug = re.sub(r"[^a-z0-9]+", "-", slug)
    slug = slug.strip("-")
    return slug


def build_crm_object(razon: str) -> dict | None:
    """Construye el objeto CRM canónico para una empresa dado su nombre de razón social.

    Retorna None si la empresa no existe en la base de datos.
    Combina datos agregados de proyectos con el estado actual en el pipeline de leads.
    """
    conn = get_db()
    try:
        cur = get_cursor(conn)

        # Fetch all project rows for this company
        cur.execute(
            _sql(
                """
                SELECT
                    codigo,
                    titulo_del_proyecto,
                    "año_adjudicacion",
                    CAST(aprobado_corfo AS REAL)  AS monto,
                    region_ejecucion,
                    sector_economico,
                    tipo_innovacion,
                    tipo_proyecto,
                    tendencia_final,
                    sostenible,
                    economia_circular_si_no
                FROM proyectos
                WHERE razon = ?
                ORDER BY "año_adjudicacion" DESC
                """
            ),
            (razon,),
        )
        rows = cur.fetchall()

        if not rows:
            return None

        # Build list of project dicts — works for both sqlite3.Row and RealDictRow
        def _row_to_dict(row) -> dict:
            if isinstance(row, dict):
                return dict(row)
            cols = [desc[0] for desc in cur.description]
            return dict(zip(cols, row))

        # Build list of project dicts
        proyectos_list: list[dict] = []
        for row in rows:
            r = _row_to_dict(row)
            proyectos_list.append({
                "codigo":            r["codigo"],
                "nombre_proyecto":   r["titulo_del_proyecto"],
                "año":               r["año_adjudicacion"],
                "monto":             r["monto"],
                "region":            r["region_ejecucion"],
                "sector":            r["sector_economico"],
                "tipo_innovacion":   r["tipo_innovacion"],
                "tipo_proyecto":     r["tipo_proyecto"],
                "tendencia":         r["tendencia_final"],
                "sostenible":        r["sostenible"] == "Sí",
                "economia_circular": r["economia_circular_si_no"] == "Sí",
            })

        # Company-level aggregations
        montos = [p["monto"] for p in proyectos_list if p["monto"] is not None]
        total_adjudicado = sum(montos) if montos else 0.0

        años = [p["año"] for p in proyectos_list if p["año"] is not None]
        primer_proyecto = min(años) if años else None
        ultimo_proyecto = max(años) if años else None

        regiones  = sorted({p["region"]  for p in proyectos_list if p["region"]})
        sectores  = sorted({p["sector"]  for p in proyectos_list if p["sector"]})
        tendencias = sorted({
            p["tendencia"] for p in proyectos_list
            if p["tendencia"] and p["tendencia"] != "Sin tendencia"
        })

        sostenible        = any(p["sostenible"]        for p in proyectos_list)
        economia_circular = any(p["economia_circular"] for p in proyectos_list)

        # Lead status — match by nombre_compania (any user)
        cur.execute(
            _sql("SELECT status FROM leads WHERE nombre_compania = ? LIMIT 1"),
            (razon,),
        )
        lead_row = cur.fetchone()
        en_leads    = lead_row is not None
        if lead_row is None:
            lead_status = None
        elif isinstance(lead_row, dict):
            lead_status = lead_row["status"]
        else:
            lead_status = lead_row[0]

    finally:
        conn.close()

    return {
        "crm_id":           _slugify(razon),
        "nombre":           razon,
        "total_adjudicado": total_adjudicado,
        "num_proyectos":    len(proyectos_list),
        "primer_proyecto":  primer_proyecto,
        "ultimo_proyecto":  ultimo_proyecto,
        "regiones":         regiones,
        "sectores":         sectores,
        "tendencias":       tendencias,
        "sostenible":       sostenible,
        "economia_circular": economia_circular,
        "en_leads":         en_leads,
        "lead_status":      lead_status,
        "proyectos":        proyectos_list,
    }


# ─────────────────────────────────────────────────────────────────────────────
# CRM FIELD MAPPING
# ─────────────────────────────────────────────────────────────────────────────

# Module-level cache: loaded once on first call to _get_crm_mapping()
_CRM_MAPPING: dict | None = None

_CRM_MAPPING_PATH = Path(__file__).parent / "context" / "crm_field_mapping.json"


def _get_crm_mapping() -> dict:
    """Carga y cachea el archivo crm_field_mapping.json. Solo lee disco en la primera llamada."""
    global _CRM_MAPPING
    if _CRM_MAPPING is None:
        with open(_CRM_MAPPING_PATH, encoding="utf-8") as f:
            _CRM_MAPPING = json.load(f)
    return _CRM_MAPPING


def apply_field_mapping(crm_object: dict, destination: str) -> dict | None:
    """Aplica el mapeo de campos al objeto CRM canónico para el destino indicado.

    Returns None if the destination is not defined in the mapping config.
    Returns a dict with keys 'destination', 'object_type', and 'properties'.
    Fields in excluded_fields or with None values in crm_object are skipped.
    The 'join_semicolon' transform joins list values with '; '.
    """
    mapping = _get_crm_mapping()
    destinations = mapping.get("destinations", {})

    if destination not in destinations:
        return None

    dest_config = destinations[destination]
    object_type: str = dest_config.get("object_type", "company")
    field_map: dict = dest_config.get("field_map", {})
    excluded_fields: list = dest_config.get("excluded_fields", [])

    properties: dict = {}
    for canonical_field, field_def in field_map.items():
        if canonical_field in excluded_fields:
            continue

        value = crm_object.get(canonical_field)
        if value is None:
            continue

        transform = field_def.get("transform")
        if transform == "join_semicolon":
            if isinstance(value, list):
                value = "; ".join(str(v) for v in value)
            else:
                value = str(value)

        property_name: str = field_def["property"]
        properties[property_name] = value

    return {
        "destination": destination,
        "object_type": object_type,
        "properties": properties,
    }


@app.route("/api/crm/mapping", methods=["GET"])
@login_required
def crm_mapping():
    """Retorna la configuración completa de mapeo de campos CRM."""
    try:
        mapping = _get_crm_mapping()
    except Exception as exc:
        log.error("Error al cargar crm_field_mapping.json: %s", exc)
        return jsonify({"error": f"Error al cargar la configuración de mapeo: {exc}"}), 500
    return jsonify(mapping), 200


@app.route("/api/crm/empresa/<path:razon>/preview/<destination>", methods=["GET"])
@login_required
def crm_empresa_preview(razon: str, destination: str):
    """Retorna el objeto CRM de una empresa transformado según el mapeo del destino indicado."""
    audit_log.info(
        "CRM_EMPRESA_PREVIEW | razon=%s | destination=%s | ip=%s",
        razon, destination, _get_client_ip(),
    )
    try:
        crm_obj = build_crm_object(razon)
    except Exception as exc:
        log.error("Error en crm_empresa_preview razon=%s: %s", razon, exc)
        return jsonify({"error": f"Error al consultar la base de datos: {exc}"}), 500

    if crm_obj is None:
        return jsonify({"error": "Empresa no encontrada"}), 404

    mapped = apply_field_mapping(crm_obj, destination)
    if mapped is None:
        return jsonify({"error": "Destino CRM no reconocido"}), 400

    return jsonify(mapped), 200


@app.route("/api/crm/empresa/<path:razon>", methods=["GET"])
@login_required
def crm_empresa(razon: str):
    """Retorna el objeto CRM canónico de una empresa para exportación o sincronización con CRM externo."""
    audit_log.info("CRM_EMPRESA | razon=%s | ip=%s", razon, _get_client_ip())
    try:
        obj = build_crm_object(razon)
    except Exception as exc:
        log.error("Error en crm_empresa razon=%s: %s", razon, exc)
        return jsonify({"error": f"Error al consultar la base de datos: {exc}"}), 500

    if obj is None:
        return jsonify({"error": "Empresa no encontrada"}), 404

    return jsonify(obj), 200


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES — SINCRONIZACIÓN
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/api/sync", methods=["POST"])
@login_required
@role_required("admin")
def sync_data():
    """Dispara sincronización manual desde datainnovacion.cl."""
    ip = _get_client_ip()
    allowed, retry_after = _sync_limiter.is_allowed(ip)
    if not allowed:
        log.warning("Rate limit superado: /api/sync desde %s", ip)
        return jsonify({
            "error": f"Límite de solicitudes excedido. Intenta de nuevo en {retry_after} segundos."
        }), 429

    try:
        result = run_sync()
        sync_status = result.get("status", "error")
        rows_fetched = result.get("rows_fetched", 0)
        audit_log.info(
            "AUDIT | action=SYNC_TRIGGERED | ip=%s | status=%s | rows_fetched=%s",
            ip, sync_status, rows_fetched,
        )
        if sync_status == "success":
            _trigger_embeddings_rebuild()
            return jsonify({
                "mensaje": "Sincronización completada exitosamente. Re-indexado de embeddings en curso.",
                "detalle": result,
            }), 200
        else:
            return jsonify({
                "error": result.get("error_message", "Error desconocido durante la sincronización"),
                "detalle": result,
            }), 500
    except Exception as e:
        log.error("Error en /api/sync: %s", e)
        audit_log.info(
            "AUDIT | action=SYNC_TRIGGERED | ip=%s | status=error | rows_fetched=0",
            ip,
        )
        return jsonify({"error": str(e), "detalle": None}), 500

# ─────────────────────────────────────────────────────────────────────────────
# ROUTES — ADMINISTRACIÓN DE USUARIOS
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_default_org(conn) -> int:
    """Obtiene o crea la organización 'default'. Retorna su id."""
    cur = get_cursor(conn)
    cur.execute(_sql("SELECT id FROM organizations WHERE slug = ? LIMIT 1"), ("default",))
    row = cur.fetchone()
    if row is not None:
        return row["id"] if isinstance(row, dict) else row[0]
    if is_postgres():
        cur.execute(
            _sql(
                "INSERT INTO organizations (name, slug, plan) VALUES (?, ?, ?) RETURNING id"
            ),
            ("Default", "default", "free"),
        )
        row = cur.fetchone()
        return row["id"] if isinstance(row, dict) else row[0]
    else:
        cur.execute(
            "INSERT INTO organizations (name, slug, plan) VALUES (?, ?, ?)",
            ("Default", "default", "free"),
        )
        return cur.lastrowid


@app.route("/api/admin/users", methods=["GET"])
@login_required
@role_required("admin")
def admin_list_users():
    conn = get_db()
    try:
        df = pd.read_sql_query(
            """
            SELECT u.id, u.username, u.role, o.slug AS org_slug, u.created_at
            FROM users u
            JOIN organizations o ON o.id = u.org_id
            ORDER BY u.created_at ASC
            """,
            conn,
        )
    except Exception as e:
        conn.close()
        log.error("admin_list_users: %s", e)
        return jsonify({"error": str(e)}), 500
    conn.close()
    records = df.where(pd.notnull(df), None).to_dict("records")
    for r in records:
        if r.get("created_at") is not None:
            r["created_at"] = str(r["created_at"])
    return jsonify(records), 200


@app.route("/api/admin/users", methods=["POST"])
@login_required
@role_required("admin")
def admin_create_user():
    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    role = (data.get("role") or "").strip()

    if not username:
        return jsonify({"error": "El nombre de usuario es requerido"}), 400
    if len(password) < 8:
        return jsonify({"error": "La contraseña debe tener al menos 8 caracteres"}), 400
    if role not in _VALID_ROLES:
        return jsonify({"error": "Rol inválido. Debe ser 'admin' o 'viewer'"}), 400

    conn = get_db()
    try:
        org_id = _ensure_default_org(conn)
        cur = get_cursor(conn)
        cur.execute(
            _sql(
                "SELECT 1 FROM users u "
                "JOIN organizations o ON o.id = u.org_id "
                "WHERE u.username = ? AND o.slug = ? LIMIT 1"
            ),
            (username, "default"),
        )
        if cur.fetchone() is not None:
            conn.close()
            return jsonify({"error": "Usuario ya existe"}), 409

        pw_hash = generate_password_hash(password)
        if is_postgres():
            cur.execute(
                _sql(
                    "INSERT INTO users (org_id, username, password_hash, role) "
                    "VALUES (?, ?, ?, ?) RETURNING id"
                ),
                (org_id, username, pw_hash, role),
            )
            row = cur.fetchone()
            new_id = row["id"] if isinstance(row, dict) else row[0]
        else:
            cur.execute(
                "INSERT INTO users (org_id, username, password_hash, role) VALUES (?, ?, ?, ?)",
                (org_id, username, pw_hash, role),
            )
            new_id = cur.lastrowid
        conn.commit()
    except Exception as e:
        conn.rollback() if is_postgres() else None
        conn.close()
        log.error("admin_create_user: %s", e)
        return jsonify({"error": str(e)}), 500
    conn.close()
    actor = session.get("username", "desconocido")
    audit_log.info(
        "AUDIT | action=USER_CREATED | actor=%s | new_user=%s | role=%s",
        actor, username, role,
    )
    return jsonify({"ok": True, "id": new_id}), 201


@app.route("/api/admin/users/<int:user_id>", methods=["PUT"])
@login_required
@role_required("admin")
def admin_update_user(user_id: int):
    data = request.get_json(silent=True) or {}
    new_role = data.get("role")
    new_password = data.get("password")

    if new_role is None and new_password is None:
        return jsonify({"error": "Se requiere al menos un campo: role o password"}), 400

    actor = session.get("username", "desconocido")

    if new_role is not None:
        if new_role not in _VALID_ROLES:
            return jsonify({"error": "Rol inválido. Debe ser 'admin' o 'viewer'"}), 400
        # Block only if the admin is actually changing their own role, not just re-sending it unchanged.
        conn = get_db()
        cur = get_cursor(conn)
        cur.execute(_sql("SELECT username, role FROM users WHERE id = ? LIMIT 1"), (user_id,))
        row = cur.fetchone()
        conn.close()
        if row is not None:
            target_username = row["username"] if isinstance(row, dict) else row[0]
            current_role = row["role"] if isinstance(row, dict) else row[1]
            if target_username == actor and new_role != current_role:
                return jsonify({"error": "No puedes modificar tu propio rol"}), 400

    if new_password is not None and len(new_password) < 8:
        return jsonify({"error": "La contraseña debe tener al menos 8 caracteres"}), 400

    conn = get_db()
    ph = "%s" if is_postgres() else "?"
    sets: list[str] = []
    params: list = []

    if new_role is not None:
        sets.append(f"role = {ph}")
        params.append(new_role)
    if new_password is not None:
        sets.append(f"password_hash = {ph}")
        params.append(generate_password_hash(new_password))

    params.append(user_id)
    try:
        cur = get_cursor(conn)
        cur.execute(f"UPDATE users SET {', '.join(sets)} WHERE id = {ph}", params)
        if cur.rowcount == 0:
            conn.close()
            return jsonify({"error": "Usuario no encontrado"}), 404
        conn.commit()
    except Exception as e:
        conn.rollback() if is_postgres() else None
        conn.close()
        log.error("admin_update_user id=%s: %s", user_id, e)
        return jsonify({"error": str(e)}), 500
    conn.close()
    audit_log.info(
        "AUDIT | action=USER_UPDATED | actor=%s | target_id=%s",
        actor, user_id,
    )
    return jsonify({"ok": True}), 200


@app.route("/api/admin/users/<int:user_id>", methods=["DELETE"])
@login_required
@role_required("admin")
def admin_delete_user(user_id: int):
    actor = session.get("username", "desconocido")

    conn = get_db()
    cur = get_cursor(conn)
    cur.execute(_sql("SELECT username FROM users WHERE id = ? LIMIT 1"), (user_id,))
    row = cur.fetchone()
    if row is None:
        conn.close()
        return jsonify({"error": "Usuario no encontrado"}), 404

    target_username = row["username"] if isinstance(row, dict) else row[0]
    if target_username == actor:
        conn.close()
        return jsonify({"error": "No puedes eliminar tu propia cuenta"}), 400

    try:
        cur.execute(_sql("DELETE FROM users WHERE id = ?"), (user_id,))
        conn.commit()
    except Exception as e:
        conn.close()
        log.error("admin_delete_user id=%s: %s", user_id, e)
        return jsonify({"error": str(e)}), 500
    conn.close()
    audit_log.info(
        "AUDIT | action=USER_DELETED | actor=%s | target_id=%s",
        actor, user_id,
    )
    return jsonify({"ok": True}), 200


@app.route("/api/admin/rebuild-embeddings", methods=["POST"])
@login_required
@role_required("admin")
def rebuild_embeddings_endpoint():
    """Trigger a manual embeddings rebuild. Returns immediately; rebuild runs in background."""
    if not _embed_rebuild_lock.acquire(blocking=False):
        return jsonify({"mensaje": "Ya hay un rebuild en curso."}), 202
    _embed_rebuild_lock.release()
    _trigger_embeddings_rebuild()
    return jsonify({"mensaje": "Rebuild de embeddings iniciado en background."}), 202


# ─────────────────────────────────────────────────────────────────────────────
# ERROR HANDLERS
# ─────────────────────────────────────────────────────────────────────────────

@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Recurso no encontrado"}), 404


@app.errorhandler(405)
def method_not_allowed(e):
    return jsonify({"error": "Método no permitido"}), 405


@app.errorhandler(Exception)
def unhandled_exception(e):
    log.error(
        "Unhandled exception | %s %s | %s | %s",
        request.method,
        request.path,
        request.remote_addr,
        "".join(traceback.format_exception(type(e), e, e.__traceback__)),
    )
    return jsonify({"error": "Error interno del servidor"}), 500


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES — ACTIVIDAD POR EMPRESA (DOB-198)
# ─────────────────────────────────────────────────────────────────────────────

ACTIVIDAD_TIPOS = ['Llamada', 'Email', 'Reunión', 'Otro']


@app.route('/api/empresa-actividad/<path:razon>', methods=['GET'])
@login_required
def get_empresa_actividad(razon: str):
    """Devuelve el registro de actividad de una empresa."""
    conn = get_db()
    try:
        df = pd.read_sql_query(
            _sql(
                "SELECT id, fecha, tipo, con_quien, nota, user_id, created_at "
                "FROM actividad "
                "WHERE nombre_compania = ? "
                "ORDER BY fecha DESC, created_at DESC "
                "LIMIT 200"
            ),
            conn,
            params=[razon],
        )
    except Exception as e:
        conn.close()
        log.error("Error en get_empresa_actividad razon=%s: %s", razon, e)
        return jsonify({'error': f'Error al consultar la base de datos: {str(e)}'}), 500
    conn.close()
    return jsonify({'actividad': df.where(pd.notnull(df), None).to_dict('records')})


@app.route('/api/empresa-actividad/<path:razon>', methods=['POST'])
@login_required
@role_required("admin")
def create_empresa_actividad(razon: str):
    """Crea una entrada de actividad para una empresa."""
    data = request.json or {}
    fecha = (data.get('fecha') or '').strip()
    tipo = (data.get('tipo') or '').strip()
    con_quien = (data.get('con_quien') or '').strip() or None
    nota = (data.get('nota') or '').strip() or None

    if not fecha:
        return jsonify({'error': 'El campo fecha es requerido'}), 400
    if not tipo:
        return jsonify({'error': 'El campo tipo es requerido'}), 400
    if tipo not in ACTIVIDAD_TIPOS:
        return jsonify({'error': f'Tipo inválido. Valores permitidos: {", ".join(ACTIVIDAD_TIPOS)}'}), 400

    user_id = session.get('username', 'default')
    conn = get_db()
    c = get_cursor(conn)
    try:
        if is_postgres():
            c.execute(
                _sql(
                    "INSERT INTO actividad (nombre_compania, fecha, tipo, con_quien, nota, user_id) "
                    "VALUES (?, ?, ?, ?, ?, ?) RETURNING id"
                ),
                [razon, fecha, tipo, con_quien, nota, user_id],
            )
            row = c.fetchone()
            new_id = row["id"] if isinstance(row, dict) else row[0]
        else:
            c.execute(
                "INSERT INTO actividad (nombre_compania, fecha, tipo, con_quien, nota, user_id) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                [razon, fecha, tipo, con_quien, nota, user_id],
            )
            new_id = c.lastrowid
        conn.commit()
    except Exception as e:
        conn.close()
        log.error("Error al crear actividad para %s: %s", razon, e)
        return jsonify({'error': str(e)}), 500

    conn.close()
    audit_log.info(
        "AUDIT | action=ACTIVIDAD_CREATED | ip=%s | user=%s | empresa=%s | tipo=%s",
        _get_client_ip(), user_id, razon, tipo,
    )
    return jsonify({'success': True, 'id': new_id}), 201


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
_migrate_leads_table()
_ensure_actividad_table()

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_ENV") == "development"
    print(f"Servidor en http://localhost:{port}")
    print(f"   SQL: {MODEL_SQL} | Explain: {MODEL_EXPLAIN}")
    app.run(host='0.0.0.0', port=port, debug=debug, use_reloader=False)
